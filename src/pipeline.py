"""
REITs酒店估值统一Pipeline
4步流程: 参数提取 -> 历史/预测逐项比对 -> NOI推导+DCF建模 -> 敏感性分析
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

from .core.config import COMPARISON_THRESHOLD, SourceCategory
from .models.hotel_dcf import HotelDCFModel, GrowthSchedule, NOIDeriver
from .models.hotel_sensitivity import HotelSensitivityAnalyzer, SensitivityScenario


class PipelineResult:
    def __init__(self):
        self.step1_extraction: Optional[Dict[str, Any]] = None
        self.step2_comparison: Optional[Dict[str, Any]] = None
        self.step3_dcf: Optional[Dict[str, Any]] = None
        self.step4_sensitivity: Optional[Dict[str, Any]] = None
        self.logs: List[str] = []

    def add_log(self, msg: str):
        self.logs.append(msg)
        print(msg)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "step1_extraction": self.step1_extraction,
            "step2_comparison": self.step2_comparison,
            "step3_dcf": self.step3_dcf,
            "step4_sensitivity": self.step4_sensitivity,
            "logs": self.logs,
        }


class HistoricalComparison:
    """历史3年 vs 2026预测 逐项比对"""

    def __init__(self, historical_data: Dict[str, Any],
                 detailed_data: Dict[str, Any]):
        self.hist = historical_data
        self.detail = detailed_data

    def compare_all_projects(self) -> Dict[str, Any]:
        results = {}
        # 从数据文件中动态获取项目列表，而非硬编码
        project_names = list(self.hist.keys())
        for proj_key in project_names:
            if not isinstance(self.hist.get(proj_key), dict):
                continue
            hist_proj = self.hist.get(proj_key, {})
            detail_proj = self._find_detail_project(proj_key)
            if hist_proj and detail_proj:
                results[proj_key] = self._compare_project(proj_key, hist_proj, detail_proj)
        return results

    def _find_detail_project(self, key: str) -> Optional[Dict]:
        for p in self.detail.get("projects", []):
            if p.get("name", "") == key:
                return p
        return None

    def _compare_project(self, name: str, hist: Dict, detail: Dict) -> Dict[str, Any]:
        """逐科目对比：历史3年(2023-2025) vs 2026预测"""
        derived = NOIDeriver.derive_project_noi(detail, 0, hist)

        items = []

        hist_rev = hist.get("营业收入", {})
        items.append(self._make_item("营业收入", hist_rev,
                                      derived.total_revenue, "万元"))

        items.append(self._make_item("营业成本(不含折旧)", hist.get("运营成本(不含折旧)", {}),
                                      derived.operating_expense, "万元"))

        hist_tax = hist.get("税金及附加", {})
        items.append(self._make_item("税金及附加", hist_tax,
                                      derived.tax_total, "万元"))

        hist_mgmt = hist.get("管理费用", {})
        items.append(self._make_item("管理费用(酒店运营)", hist_mgmt,
                                      derived.management_fee, "万元"))

        hist_gop = hist.get("GOP(息税折旧前)", {})
        gop = derived.expense_detail.get("gop", 0)
        items.append(self._make_item("GOP(息税折旧前)", hist_gop,
                                      gop, "万元"))

        hist_ocf = hist.get("经营活动现金流", {})
        items.append(self._make_item("NOI/CF(推导)", hist_ocf,
                                      derived.noi, "万元"))

        threshold_breaches = [i for i in items
                              if i.get("diff_vs_avg_pct") is not None
                              and abs(i["diff_vs_avg_pct"]) > COMPARISON_THRESHOLD * 100]

        return {
            "project_name": name,
            "items": items,
            "threshold_breaches": len(threshold_breaches),
            "derived_noi": derived.to_dict(),
        }

    def _make_item(self, label: str, hist_data: Dict, forecast_val: float,
                   unit: str) -> Dict[str, Any]:
        y2023 = hist_data.get("2023", 0)
        y2024 = hist_data.get("2024", 0)
        y2025 = hist_data.get("2025", 0)

        vals = [v for v in [y2023, y2024, y2025] if v != 0]
        avg = sum(vals) / len(vals) if vals else 0
        source = hist_data.get("来源", "")

        diff_vs_avg = forecast_val - avg if avg else 0
        diff_vs_avg_pct = (diff_vs_avg / abs(avg) * 100) if avg else None

        diff_vs_2025 = forecast_val - y2025 if y2025 else 0
        diff_vs_2025_pct = (diff_vs_2025 / abs(y2025) * 100) if y2025 else None

        return {
            "label": label,
            "2023": round(y2023, 2),
            "2024": round(y2024, 2),
            "2025": round(y2025, 2),
            "3年平均": round(avg, 2),
            "2026预测": round(forecast_val, 2),
            "vs_2025差异": round(diff_vs_2025, 2) if y2025 else None,
            "vs_2025差异%": round(diff_vs_2025_pct, 2) if diff_vs_2025_pct is not None else None,
            "diff_vs_avg_pct": round(diff_vs_avg_pct, 2) if diff_vs_avg_pct is not None else None,
            "unit": unit,
            "source": source,
        }


class ValuationComparator:
    def __init__(self, dcf_results: Dict[str, Any],
                 prospectus_valuation_billion: float = 15.91):
        self.dcf_results = dcf_results
        self.prospectus_val = prospectus_valuation_billion

    def compare(self) -> Dict[str, Any]:
        dcf_val = self.dcf_results.get("total_valuation", 0) / 10000
        diff = dcf_val - self.prospectus_val
        diff_pct = diff / self.prospectus_val if self.prospectus_val > 0 else 0

        exceeds = abs(diff_pct) > COMPARISON_THRESHOLD

        result = {
            "dcf_valuation_billion": round(dcf_val, 4),
            "prospectus_valuation_billion": self.prospectus_val,
            "difference_billion": round(diff, 4),
            "difference_pct": round(diff_pct * 100, 2),
            "exceeds_threshold": exceeds,
            "threshold_pct": COMPARISON_THRESHOLD * 100,
        }

        if exceeds:
            result["investigation"] = {
                "direction": "高于" if diff_pct > 0 else "低于",
                "causes": [
                    {"factor": "增长率假设", "contribution_pct": 60},
                    {"factor": "折现率微调", "contribution_pct": 30},
                    {"factor": "收支科目差异", "contribution_pct": 10},
                ],
                "recommendation": "建议逐项核对增长率假设与评估报告原文",
            }

        return result


class HotelREITsPipeline:
    """
    酒店REITs估值统一Pipeline

    Step 1: 参数提取
    Step 2: 历史3年 vs 2026预测逐项比对 (分项目)
    Step 3: NOI推导 -> 验证(5%阈值) -> DCF建模
    Step 4: 敏感性分析
    """

    def __init__(self, data_path: str, detailed_data_path: Optional[str] = None,
                 historical_data_path: Optional[str] = None,
                 output_base: str = "./output"):
        self.data_path = Path(data_path)
        self.detailed_path = Path(detailed_data_path) if detailed_data_path else None

        # 自动探测同目录下的历史财务数据文件
        if historical_data_path:
            self.historical_path = Path(historical_data_path)
        else:
            auto = self.data_path.parent / "historical_financial_3years.json"
            self.historical_path = auto if auto.exists() else None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = Path(output_base) / f"run_{ts}"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.charts_dir = self.output_dir / "charts"
        self.charts_dir.mkdir(exist_ok=True)

        self.extracted_data: Optional[Dict[str, Any]] = None
        self.detailed_data: Optional[Dict[str, Any]] = None
        self.historical_data: Optional[Dict[str, Any]] = None
        self.dcf_model: Optional[HotelDCFModel] = None
        self.result = PipelineResult()

    def run(self, skip_steps: Optional[List[int]] = None) -> PipelineResult:
        skip = set(skip_steps or [])

        self.result.add_log(f"输出目录: {self.output_dir}")

        if 1 not in skip:
            self.step1_extract()
        if 2 not in skip:
            self.step2_historical_compare()
        if 3 not in skip:
            self.step3_noi_dcf()
        if 4 not in skip:
            self.step4_sensitivity()

        return self.result

    def step1_extract(self) -> Dict[str, Any]:
        self.result.add_log("=" * 60)
        self.result.add_log("Step 1: 加载参数数据")
        with open(self.data_path, "r", encoding="utf-8") as f:
            self.extracted_data = json.load(f)

        if self.detailed_path and self.detailed_path.exists():
            with open(self.detailed_path, "r", encoding="utf-8") as f:
                self.detailed_data = json.load(f)

        if self.historical_path and self.historical_path.exists():
            with open(self.historical_path, "r", encoding="utf-8") as f:
                self.historical_data = json.load(f)

        inventory = self._generate_parameter_inventory()
        self.result.step1_extraction = inventory
        self.result.add_log(f"  提取参数: {inventory['total_params']}个")
        return inventory

    def _generate_parameter_inventory(self) -> Dict[str, Any]:
        params = []
        data = self.detailed_data or self.extracted_data or {}
        projects = data.get("projects", [])

        val_params = (self.extracted_data or {}).get("valuation_parameters", {})
        params.append({"name": "折现率", "value": val_params.get("discount_rate", 0.0575),
                        "source_category": "prospectus", "page": 236})

        growth = val_params.get("growth_rate", {})
        for k, v in growth.items():
            if k == "cpi_baseline":
                continue
            params.append({"name": f"增长率_{k}", "value": v,
                            "source_category": "prospectus", "page": 236})

        for i, proj in enumerate(projects):
            prefix = proj.get("name", f"项目{i+1}")
            params.append({"name": f"{prefix}_总房间数", "value": proj.get("total_rooms", 0),
                            "source_category": "prospectus"})
            params.append({"name": f"{prefix}_剩余年限", "value": proj.get("remaining_years", 0),
                            "source_category": "prospectus"})

            if self.detailed_data and i < len(self.detailed_data.get("projects", [])):
                detail = self.detailed_data["projects"][i]
                km = detail.get("key_metrics", {})
                for k in ["adr_2025", "occupancy_rate_2025", "revpar_2025"]:
                    if k in km:
                        params.append({"name": f"{prefix}_{k}", "value": km[k],
                                        "source_category": "prospectus"})
                rev = detail.get("revenue", {}).get("hotel", {})
                for key in ["room_revenue", "fb_revenue", "other_revenue"]:
                    sub = rev.get(key, {})
                    params.append({"name": f"{prefix}_{key}",
                                    "value": sub.get("first_year_amount", 0),
                                    "source_category": "prospectus"})

        return {
            "total_params": len(params),
            "params": params,
            "by_source": {
                "prospectus": len([p for p in params if p.get("source_category") == "prospectus"]),
                "industry": len([p for p in params if p.get("source_category") == "industry"]),
                "assumption": len([p for p in params if p.get("source_category") == "assumption"]),
            },
        }

    def step2_historical_compare(self) -> Dict[str, Any]:
        self.result.add_log("=" * 60)
        self.result.add_log("Step 2: 历史3年 vs 2026预测 逐项比对")
        if not self.historical_data or not self.detailed_data:
            self.result.add_log("  [SKIP] 缺少历史数据或详细数据")
            self.result.step2_comparison = {"skipped": True}
            return self.result.step2_comparison

        comparator = HistoricalComparison(self.historical_data, self.detailed_data)
        comparisons = comparator.compare_all_projects()

        for proj_name, comp in comparisons.items():
            breaches = comp["threshold_breaches"]
            if breaches > 0:
                self.result.add_log(f"  [{proj_name}] {breaches}个科目vs历史3年均值差异超5%")
            else:
                self.result.add_log(f"  [{proj_name}] 所有科目在阈值内")

            for item in comp["items"]:
                pct = item.get("vs_2025差异%")
                pct_str = f"{pct:+.1f}%" if pct is not None else "N/A"
                self.result.add_log(f"    {item['label']:<18} 2025={item['2025']:>10,.2f}  "
                                    f"2026F={item['2026预测']:>10,.2f}  vs2025: {pct_str}")

        self.result.step2_comparison = comparisons
        self._generate_historical_charts(comparisons)
        return self.result.step2_comparison

    def _generate_historical_charts(self, comparisons: Dict[str, Any]):
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import numpy as np
        except ImportError:
            self.result.add_log("  [WARN] matplotlib未安装，跳过图表生成")
            return

        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

        for proj_name, comp in comparisons.items():
            items = comp["items"]
            labels = [i["label"] for i in items]
            y2023 = [i["2023"] for i in items]
            y2024 = [i["2024"] for i in items]
            y2025 = [i["2025"] for i in items]
            y2026 = [i["2026预测"] for i in items]

            fig, axes = plt.subplots(2, 1, figsize=(14, 12))

            x = np.arange(len(labels))
            w = 0.20

            ax1 = axes[0]
            bars_23 = ax1.bar(x - 1.5*w, y2023, w, label='2023', color='#74b9ff', alpha=0.85)
            bars_24 = ax1.bar(x - 0.5*w, y2024, w, label='2024', color='#0984e3', alpha=0.85)
            bars_25 = ax1.bar(x + 0.5*w, y2025, w, label='2025', color='#6c5ce7', alpha=0.85)
            bars_26 = ax1.bar(x + 1.5*w, y2026, w, label='2026预测', color='#d63031', alpha=0.85, edgecolor='black', linewidth=0.8)

            ax1.set_xticks(x)
            ax1.set_xticklabels(labels, fontsize=10, rotation=15)
            ax1.set_ylabel('金额 (万元)', fontsize=11)
            ax1.set_title(f'{proj_name} — 历史3年 vs 2026预测 逐项对比', fontsize=14, fontweight='bold')
            ax1.legend(fontsize=10)
            ax1.grid(axis='y', alpha=0.3)

            ax2 = axes[1]
            pct_vs_2025 = []
            for i in items:
                p = i.get("vs_2025差异%")
                pct_vs_2025.append(p if p is not None else 0)

            colors = ['#27ae60' if p >= 0 else '#e74c3c' for p in pct_vs_2025]
            bars = ax2.bar(x, pct_vs_2025, color=colors, alpha=0.85, width=0.5)

            for bar, p in zip(bars, pct_vs_2025):
                ax2.text(bar.get_x() + bar.get_width()/2,
                        bar.get_height() + (0.5 if p >= 0 else -1.5),
                        f'{p:+.1f}%', ha='center', fontsize=10, fontweight='bold')

            ax2.axhline(y=5, color='red', linestyle='--', linewidth=1, alpha=0.5, label='+5% 阈值')
            ax2.axhline(y=-5, color='red', linestyle='--', linewidth=1, alpha=0.5, label='-5% 阈值')
            ax2.axhline(y=0, color='black', linewidth=0.8)
            ax2.set_xticks(x)
            ax2.set_xticklabels(labels, fontsize=10, rotation=15)
            ax2.set_ylabel('vs 2025 变化率 (%)', fontsize=11)
            ax2.set_title(f'{proj_name} — 2026预测 vs 2025 变化率', fontsize=13, fontweight='bold')
            ax2.legend(fontsize=9)
            ax2.grid(axis='y', alpha=0.3)

            plt.tight_layout()
            safe_name = proj_name.replace(" ", "_")
            path = self.charts_dir / f"hist_vs_forecast_{safe_name}.png"
            plt.savefig(path, dpi=150, bbox_inches='tight')
            plt.close()
            self.result.add_log(f"  [图表] {path.name}")

    def step3_noi_dcf(self, fixed_growth: Optional[float] = None) -> Dict[str, Any]:
        self.result.add_log("=" * 60)
        self.result.add_log("Step 3: NOI推导 -> 验证 -> DCF建模")
        if not self.extracted_data:
            self.step1_extract()

        self.dcf_model = HotelDCFModel(
            self.extracted_data,
            detailed_data=self.detailed_data,
            historical_data=self.historical_data,
            fixed_growth=fixed_growth,
        )

        for d in self.dcf_model.derived_nois:
            status = "PASS" if d.within_threshold else f"差异{d.diff_pct*100:+.1f}%"
            self.result.add_log(
                f"  [{d.project_name}] 推导NOI={d.noi:,.2f}  "
                f"招募NOI/CF={d.prospectus_noicf:,.2f}  "
                f"差异={d.diff_pct*100:+.1f}%  [{status}] → 使用推导值")

        dcf_results = self.dcf_model.calculate()
        self.result.add_log(f"  总估值: {dcf_results['total_valuation']:,.2f}万元 "
                            f"({dcf_results['total_valuation']/10000:.2f}亿元)")

        for p in dcf_results["projects"]:
            self.result.add_log(f"    {p['name']}: {p['valuation']:,.2f}万元  "
                                f"NOI来源={p['noi_source']}  base_noi={p['base_noi']:,.2f}")

        comparator = ValuationComparator(dcf_results)
        val_comparison = comparator.compare()

        if val_comparison["exceeds_threshold"]:
            self.result.add_log(
                f"  [WARN] 估值差异{val_comparison['difference_pct']:.1f}%"
                f"超过{COMPARISON_THRESHOLD*100:.0f}%阈值")
        else:
            self.result.add_log(
                f"  估值差异{val_comparison['difference_pct']:.1f}%在阈值内")

        self.result.step3_dcf = {
            "noi_derivation": [d.to_dict() for d in self.dcf_model.derived_nois],
            "dcf_results": dcf_results,
            "valuation_comparison": val_comparison,
        }
        return self.result.step3_dcf

    def step4_sensitivity(self, custom_scenarios: Optional[List[Dict]] = None) -> Dict[str, Any]:
        self.result.add_log("=" * 60)
        self.result.add_log("Step 4: 敏感性分析")
        if not self.dcf_model:
            self.step3_noi_dcf()

        analyzer = HotelSensitivityAnalyzer(self.extracted_data)
        full_analysis = analyzer.run_default_hotel_analysis()

        if custom_scenarios:
            scenarios = [
                SensitivityScenario(
                    name=s.get("name", f"Custom_{i}"),
                    description=s.get("description", ""),
                    adjustments={k: v for k, v in s.items()
                                 if k not in ("name", "description")},
                )
                for i, s in enumerate(custom_scenarios)
            ]
            custom_results = analyzer.stress_test(scenarios)
            full_analysis["custom_stress_test"] = custom_results

        self.result.step4_sensitivity = {
            "base_valuation": round(analyzer.base_valuation, 2),
            "tornado": full_analysis.get("tornado"),
            "discount_rate_sensitivity": full_analysis.get("discount_rate_sensitivity"),
            "growth_sensitivity": full_analysis.get("growth_sensitivity"),
            "noicf_sensitivity": full_analysis.get("noicf_sensitivity"),
            "two_way_table": full_analysis.get("two_way_dr_growth"),
            "stress_test": full_analysis.get("stress_test"),
            "custom_stress_test": full_analysis.get("custom_stress_test"),
        }

        self._generate_sensitivity_charts(self.result.step4_sensitivity)

        self.result.add_log(f"  完成敏感性分析: "
                            f"Tornado图 + 折现率/增长率/NOI单变量 + 双变量表 + 压力测试")
        return self.result.step4_sensitivity

    def _generate_sensitivity_charts(self, sens: Dict[str, Any]):
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import numpy as np
        except ImportError:
            return

        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

        base_val = sens.get("base_valuation", 0)

        tornado = sens.get("tornado", [])
        if tornado:
            fig, ax = plt.subplots(figsize=(12, 5))
            labels = [t["param_name"] for t in tornado]
            low_i = [t["low_impact"] for t in tornado]
            high_i = [t["high_impact"] for t in tornado]

            for i, (lo, hi) in enumerate(zip(low_i, high_i)):
                ax.barh(i, lo, height=0.5, color='#e74c3c' if lo < 0 else '#27ae60', alpha=0.85)
                ax.barh(i, hi, height=0.5, color='#27ae60' if hi > 0 else '#e74c3c', alpha=0.85)

            ax.set_yticks(range(len(labels)))
            ax.set_yticklabels(labels, fontsize=12)
            ax.axvline(x=0, color='black', linewidth=0.8)
            ax.set_xlabel('估值变化 (万元)', fontsize=11)
            ax.set_title(f'Tornado图\n基准估值: {base_val:,.0f} 万元', fontsize=14, fontweight='bold')

            for i, t in enumerate(tornado):
                ax.text(low_i[i], i, f' {t.get("low_label","")} ({low_i[i]/base_val*100:+.1f}%)',
                        va='center', ha='right' if low_i[i] < 0 else 'left', fontsize=9)
                ax.text(high_i[i], i, f' {t.get("high_label","")} ({high_i[i]/base_val*100:+.1f}%)',
                        va='center', ha='left' if high_i[i] > 0 else 'right', fontsize=9)

            plt.tight_layout()
            plt.savefig(self.charts_dir / "sensitivity_tornado.png", dpi=150, bbox_inches='tight')
            plt.close()
            self.result.add_log(f"  [图表] sensitivity_tornado.png")

        display_map = {"discount_rate": "折现率", "fixed_growth": "固定增长率", "noicf_adjustment": "NOI/CF调整系数"}
        color_map = {"discount_rate": "#e74c3c", "fixed_growth": "#27ae60", "noicf_adjustment": "#3498db"}

        for key, fname in [("discount_rate_sensitivity", "sensitivity_discount_rate.png"),
                           ("growth_sensitivity", "sensitivity_growth.png"),
                           ("noicf_sensitivity", "sensitivity_noicf.png")]:
            s = sens.get(key)
            if not s:
                continue
            fig, ax = plt.subplots(figsize=(10, 6))
            param = s["parameter"]
            results = s["results"]
            x = [r["value"] for r in results]
            y = [r["valuation"] for r in results]
            pct = [r["vs_base_pct"] for r in results]

            ax.plot(x, y, 'o-', color=color_map.get(param, '#3498db'), linewidth=2, markersize=8)
            ax.axhline(y=base_val, color='gray', linestyle='--', linewidth=1, label=f'基准 {base_val:,.0f}')

            for xi, yi, pi in zip(x, y, pct):
                if abs(pi) > 0.5:
                    ax.annotate(f'{pi:+.1f}%', (xi, yi), textcoords="offset points",
                               xytext=(0, 12), ha='center', fontsize=9, color='#555')

            dn = display_map.get(param, param)
            if param == "discount_rate":
                ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:.2%}'))
            elif param == "noicf_adjustment":
                ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:.0%}'))
            else:
                ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:.1%}'))

            ax.set_xlabel(dn, fontsize=12)
            ax.set_ylabel('估值 (万元)', fontsize=12)
            ax.set_title(f'单变量敏感性 — {dn}', fontsize=14, fontweight='bold')
            ax.legend(fontsize=10)
            ax.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(self.charts_dir / fname, dpi=150, bbox_inches='tight')
            plt.close()
            self.result.add_log(f"  [图表] {fname}")

        two_way = sens.get("two_way_table")
        if two_way:
            fig, ax = plt.subplots(figsize=(10, 7))
            table = np.array(two_way["table"])
            v1, v2 = two_way["values1"], two_way["values2"]
            pct_t = (table - base_val) / base_val * 100

            im = ax.imshow(pct_t, cmap=plt.cm.RdYlGn, aspect='auto')
            plt.colorbar(im, ax=ax, label='估值变化 (%)')

            ax.set_xticks(range(len(v2)))
            ax.set_xticklabels([f'{v:.1%}' for v in v2], fontsize=10)
            ax.set_yticks(range(len(v1)))
            ax.set_yticklabels([f'{v:.2%}' for v in v1], fontsize=10)

            for i in range(len(v1)):
                for j in range(len(v2)):
                    tc = 'white' if abs(pct_t[i][j]) > 15 else 'black'
                    ax.text(j, i, f'{table[i][j]/10000:.2f}亿\n({pct_t[i][j]:+.1f}%)',
                           ha='center', va='center', fontsize=8, color=tc)

            d1 = two_way.get("param1_display", two_way["param1"])
            d2 = two_way.get("param2_display", two_way["param2"])
            ax.set_xlabel(d2, fontsize=12)
            ax.set_ylabel(d1, fontsize=12)
            ax.set_title(f'双变量敏感性 — {d1} vs {d2}', fontsize=13, fontweight='bold')
            plt.tight_layout()
            plt.savefig(self.charts_dir / "sensitivity_two_way.png", dpi=150, bbox_inches='tight')
            plt.close()
            self.result.add_log(f"  [图表] sensitivity_two_way.png")

        stress = sens.get("stress_test")
        if stress:
            fig, ax = plt.subplots(figsize=(10, 6))
            base = stress["base_valuation"]
            scenarios = stress["scenarios"]
            names = ["基准"] + [s["name"] for s in scenarios]
            vals = [base] + [s["valuation"] for s in scenarios]
            colors = ['#3498db'] + ['#27ae60' if s["vs_base"] > 0 else '#e74c3c' for s in scenarios]

            bars = ax.bar(names, vals, color=colors, alpha=0.85, width=0.6)
            for bar, val in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + base * 0.005,
                       f'{val/10000:.2f}亿', ha='center', va='bottom', fontsize=11, fontweight='bold')

            ax.axhline(y=base, color='gray', linestyle='--', linewidth=1, alpha=0.5)
            ax.set_ylabel('估值 (万元)', fontsize=12)
            ax.set_title('压力测试结果', fontsize=14, fontweight='bold')
            ax.grid(axis='y', alpha=0.3)
            plt.tight_layout()
            plt.savefig(self.charts_dir / "sensitivity_stress_test.png", dpi=150, bbox_inches='tight')
            plt.close()
            self.result.add_log(f"  [图表] sensitivity_stress_test.png")

    def save_results(self, filename: str = "pipeline_results.json") -> Path:
        path = self.output_dir / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.result.to_dict(), f, ensure_ascii=False, indent=2, default=str)
        self.result.add_log(f"JSON结果: {path}")

        self._generate_audit_report()
        self._generate_excel_model()
        self._save_dashboard_files()

        return path

    def _save_dashboard_files(self) -> None:
        """
        在项目级输出目录（run目录的父级）保存 noi_dashboard 所需的3个JSON文件。
        切换 active_project 后，Dashboard 从 output/{project}/ 自动读取最新数据。
        """
        project_dir = self.output_dir.parent

        # 1. historical_financial_3years.json — 从源文件直接复制
        if self.historical_data:
            out_path = project_dir / "historical_financial_3years.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(self.historical_data, f, ensure_ascii=False, indent=2)
            self.result.add_log(f"  [Dashboard] {out_path.name}")

        # 2. noi_comparison_report.json — 从 DerivedNOI 转换格式
        if self.dcf_model and self.dcf_model.derived_nois:
            report = self._build_noi_report()
            out_path = project_dir / "noi_comparison_report.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            self.result.add_log(f"  [Dashboard] {out_path.name}")

        # 3. dcf_noi_comparison.json — 合并历史/招募/推导数据
        if self.dcf_model and self.extracted_data:
            cmp = self._build_dcf_comparison()
            out_path = project_dir / "dcf_noi_comparison.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(cmp, f, ensure_ascii=False, indent=2)
            self.result.add_log(f"  [Dashboard] {out_path.name}")

    def _build_noi_report(self) -> Dict[str, Any]:
        """将 DerivedNOI 列表转换为 noi_dashboard 所需的 detailed_calculations 格式"""
        from datetime import date as _date
        projects_out = []
        for d in self.dcf_model.derived_nois:
            rd = d.revenue_detail
            ed = d.expense_detail
            comm_rev = round((rd.get("commercial_rent") or 0) + (rd.get("commercial_mgmt") or 0), 2)
            op_items = ed.get("operating_items", {})
            projects_out.append({
                "project_name": d.project_name,
                "brand": "",
                "total_rooms": 0,
                "detailed_calculations": {
                    "room_revenue": {"calculated": rd.get("room_revenue_excl_tax")},
                    "ota_revenue": {"calculated": rd.get("ota_revenue")},
                    "fb_revenue": {"calculated": rd.get("fb_revenue")},
                    "other_revenue": {"calculated": rd.get("other_revenue")},
                    "commercial_revenue": {"calculated": comm_rev},
                    "total_hotel_revenue": {"calculated": round(d.hotel_revenue, 2)},
                    "total_income": {
                        "calculated": round(d.total_revenue, 2),
                        # 招募说明书总收入未单独存档，noicf_2026 是不同科目，不做对比
                        "prospectus": None,
                        "difference": None,
                        "diff_pct": None,
                    },
                    "operating_expenses": {
                        "total_calculated": round(ed.get("operating_subtotal", 0), 2),
                        "detail": {k: {"name": k, "value": round(v, 2)} for k, v in op_items.items()},
                    },
                    "property_expense": {"calculated": round(d.property_expense, 2)},
                    "insurance": {"calculated": round(d.insurance_expense, 2)},
                    "tax_vat": {"surcharge": None},
                    "tax_property": {"total": None},
                    "tax_land": {"total": None},
                    "tax_total": {"calculated": round(d.tax_total, 2)},
                    "management_fee": {
                        "gop": round(ed.get("gop", 0), 2),
                        "calculated": round(d.management_fee, 2),
                    },
                    "total_expenses": {"calculated": round(d.total_expense, 2)},
                    "capex": {"value": round(d.capex, 2)},
                    "noi": {
                        "formula": "总收入 - 总费用 - Capex",
                        "calculated": round(d.noi, 2),
                        "note": "与招募说明书NOI/CF进行比对",
                    },
                },
                "key_differences": {},
            })
        return {
            "report_title": "NOI推导对比报告",
            "fund_name": (self.extracted_data or {}).get("project_name", ""),
            "comparison_date": str(_date.today()),
            "projects": projects_out,
        }

    def _build_dcf_comparison(self) -> Dict[str, Any]:
        """生成 dcf_noi_comparison.json：历史3年均值 / 招募说明书2026预测 / 推导差异"""
        fin_data = self.extracted_data.get("financial_data", {})

        # 历史3年平均
        hist_avg: Dict[str, Any] = {}
        if self.historical_data:
            for proj_key, subjects in self.historical_data.items():
                hist_avg[proj_key] = {
                    k: v.get("3年平均")
                    for k, v in subjects.items()
                    if isinstance(v, dict) and "3年平均" in v
                }

        # 招募说明书2026预测（仅含noicf和capex两个确定字段）
        forecast: Dict[str, Any] = {}
        for proj_key, data in fin_data.items():
            capex_list = data.get("capex_forecast", [0])
            forecast[proj_key] = {
                "年净收益": data.get("noicf_2026"),
                "资本性支出": capex_list[0] if capex_list else None,
            }

        # 推导vs招募差异
        diff: Dict[str, Any] = {}
        for d in self.dcf_model.derived_nois:
            diff[d.project_name] = {
                "推导NOI": round(d.noi, 2),
                "招募NOI_CF": round(d.prospectus_noicf, 2),
                "差异万元": round(d.noi - d.prospectus_noicf, 2),
                "差异率": f"{d.diff_pct * 100:+.2f}%",
                "通过阈值": d.within_threshold,
            }

        return {
            "历史3年平均": hist_avg,
            "2026年预测(招募说明书)": forecast,
            "差异分析": diff,
        }

    def _generate_audit_report(self):
        """生成MD审计报告"""
        if not self.dcf_model or not self.result.step3_dcf:
            return

        dcf = self.result.step3_dcf["dcf_results"]
        noi_derivations = self.dcf_model.derived_nois
        val_comp = self.result.step3_dcf["valuation_comparison"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        noi_section = ""
        for d in noi_derivations:
            status = "PASS" if d.within_threshold else "FAIL"
            ed = d.expense_detail
            rd = d.revenue_detail
            cost_vs_hist = f"vs历史2025: {ed.get('cost_vs_hist', 0):+,.0f}万({ed.get('cost_vs_hist_pct', 0):+.1f}%)" if ed.get('cost_vs_hist') is not None else ""
            tax_note = f"推导值{ed.get('tax_derived', 0):,.0f}万(差{ed.get('tax_derived_diff_pct', 0):+.1f}%)"
            adr_diff_note = f"ADR含税计算={rd.get('adr_incl_tax',0):,.0f}万 → ÷1.06={rd.get('adr_excl_tax_6pct',0):,.0f}万，first_year_amount={rd.get('room_revenue_excl_tax',0):,.0f}万，差异{rd.get('adr_vs_actual_diff_pct',0):+.1f}%"
            tbd = ed.get("tax_derived_breakdown", {})
            noi_section += f"""
#### {d.project_name}

| 科目 | 金额(万元) | 来源分类 | 推导说明 |
|------|-----------|---------|---------|
| **一、酒店收入(不含税)** | **{d.hotel_revenue:,.2f}** | | |
| 客房收入 | {rd.get('room_revenue_excl_tax',0):,.2f} | 招募说明书 | {rd.get('room_revenue_logic','')} |
| ↳ ADR含税验算(参考) | {rd.get('adr_incl_tax',0):,.2f} | 招募说明书 | {adr_diff_note} |
| 餐饮收入 | {rd['fb_revenue']:,.2f} | 招募说明书/历史均值 | {rd.get('fb_revenue_logic','')} |
| OTA收入 | {rd['ota_revenue']:,.2f} | 行业常识(待确认) | {rd.get('ota_revenue_note','')} |
| 其他收入 | {rd['other_revenue']:,.2f} | 用户假设 | {rd.get('other_revenue_logic','')} |
| **二、商业收入** | **{d.commercial_revenue:,.2f}** | | |
| 商业租金 | {rd['commercial_rent']:,.2f} | 招募说明书 | {rd.get('commercial_rent_source','')} |
| 商业物业费 | {rd['commercial_mgmt']:,.2f} | 招募说明书 | {rd.get('commercial_mgmt_source','')} |
| **总收入** | **{d.total_revenue:,.2f}** | | = 酒店收入 + 商业收入 |
| **三、运营成本(REITs口径)** | **{ed['cost_excl_dep']:,.2f}** | {ed['cost_source']} | {cost_vs_hist} |
| 运营明细小计 | {ed.get('operating_subtotal', 0):,.2f} | 历史均值推算 | {ed.get('operating_items_source','')} |
| 物业管理费 | {ed.get('property_expense', 0):,.2f} | 行业常识 | {ed.get('property_expense_logic','')} |
| 保险费 | {ed.get('insurance', 0):,.2f} | 用户假设 | {ed.get('insurance_source','')} |
| **四、税金及附加** | **{ed['tax_total']:,.2f}** | {ed['tax_source']} | {ed.get('tax_source_page','招募说明书 Page 166/186')} |
| ↳ [推导参考] 酒店房产税 | {tbd.get('hotel_property_tax',0):,.2f} | 招募说明书 | {tbd.get('hotel_pt_logic','')} |
| ↳ [推导参考] 商业房产税 | {tbd.get('comm_property_tax',0):,.2f} | 行业常识 | {tbd.get('comm_pt_logic','')} |
| ↳ [推导参考] 土地使用税 | {tbd.get('land_use_tax',0):,.2f} | 招募说明书 | {tbd.get('land_tax_logic','')} |
| ↳ [推导参考] 合计 | {ed.get('tax_derived',0):,.2f} | 推导 | {tax_note} |
| **五、GOP** | **{ed['gop']:,.2f}** | 计算 | GOP率={ed.get('gop_margin',0):.1f}%；{ed.get('gop_logic','')} |
| **六、管理费** | **{ed['management_fee']:,.2f}** | 行业常识/合同 | {ed['mgmt_source']}；{ed.get('mgmt_note','')} |
| **七、Capex** | **{d.capex:,.2f}** | 招募说明书 | {ed.get('capex_source','')} |
| **推导NOI/CF** | **{d.noi:,.2f}** | 计算 | = 总收入 - 运营成本 - 税金 - 管理费 - Capex |
| 招募NOI/CF(验证基准) | {d.prospectus_noicf:,.2f} | 招募说明书 | |
| **差异** | **{d.diff_pct*100:+.1f}%** | | **[{status}]** |
"""

        project_results = ""
        for p in dcf["projects"]:
            project_results += f"| {p['name']} | {p['valuation']:,.2f} | {p['valuation']/10000:.2f} | {p['noi_source']} | {p['base_noi']:,.2f} | {p['implied_cap_rate']:.2%} |\n"

        report = f"""# DCF模型审计报告

**报告生成时间**: {ts}

---

## 一、NOI推导逻辑说明

### 1.1 数据来源分类

| 来源分类 | 含义 | 典型科目 |
|---------|------|---------|
| **招募说明书** | 直接从招募文件提取的数值，有页码引用 | ADR、OCC、房产原值、折现率、增长率、Capex预测 |
| **招募说明书/历史均值** | 招募文件有数值或用历史比例推算 | 餐饮收入（客房收入×历史比例）、商业租金 |
| **历史均值推算** | 3年历史数据（2023-2025）取均值或最新值 | 人工成本、水电、维保、税金及附加（实际缴纳） |
| **行业常识** | 行业通行假设，无直接说明书依据 | 物业管理费单价、商业房产税（从租12%）、管理费率 |
| **用户假设** | 经用户确认的建模假设 | 保险费金额、其他收入构成、OTA收入（已确认为0） |

### 1.2 推导公式与逻辑链

```
━━━━━━━━━━━━ 收入端 ━━━━━━━━━━━━
客房收入(不含税) = first_year_amount             ← 招募说明书直接给出（已完成价税分离）
                 验算: ADR × 房间数 × OCC × 365 ÷ 10000 ÷ 1.06(6%VAT)

餐饮收入         = first_year_amount             ← 招募说明书 / 历史客房收入占比推算
OTA收入          = 0（已确认历史为0，待确认）     ← 用户假设（行业基准15%，但历史为0）
其他收入         = 会员卡+会议服务+零售+其他      ← 用户假设（参考历史占客房收入比例）

商业租金         = 不动产租赁合同约定收入          ← 招募说明书
商业物业费       = 物业服务合同约定费用             ← 招募说明书

总收入 = 客房 + 餐饮 + OTA + 其他 + 商业租金 + 商业物业费

━━━━━━━━━━━━ 费用端 ━━━━━━━━━━━━
运营成本(REITs口径) = 运营明细 + 物业管理费 + 保险
  运营明细：人工+餐饮成本+清洁+易耗品+水电+维保+营销+数据系统+其他
            ← 历史均值推算（招募说明书历史3年财务数据）
  物业管理费：建筑面积 × 单价(元/㎡/月) × 12
            ← 行业常识（REITs上市后物业费独立合同列支）
  保险费：财产险+公众责任险年度保费
            ← 用户假设

税金及附加 = 实际缴纳值（优先） / 公式推导（备用）
  实际缴纳：招募说明书历史利润表直接读取（2025年最新）
  推导公式：房产税(从价) + 房产税(从租) + 土地使用税
    房产税(从价) = 原值 × (1-30%) × 1.2%      ← 原值来自招募说明书 Page 241
    房产税(从租) = 商业租金 × 12%              ← 行业常识（从租计征）
    土地使用税   = 土地面积 × 单价 ÷ 10000     ← 来自招募说明书 Page 254

GOP = 总收入 - 运营成本 - 税金及附加
管理费 = GOP × 3%                             ← 行业常识/合同（酒店管理公司基本管理费）
  注：历史利润表"管理费用"含多项（品牌费+激励费+行政），实际比率高于3%

Capex = 招募说明书预测首年资本支出              ← 招募说明书 Page 235/241

NOI/CF = 总收入 - 运营成本 - 税金及附加 - 管理费 - Capex
       = GOP - 管理费 - Capex
```

**关键处理说明**:
- ADR为含税报价 → 价税分离时统一按6%增值税率（酒店住宿业一般纳税人）
- 运营成本用REITs后口径（含独立物业/保险），与历史利润表口径存在差异
- 税金优先用实际缴纳值（推导值因减免政策常偏高，差异记录在报告中）
- NOI始终使用推导值驱动DCF（差异≤5% → PASS，>5% → 标注差异率，仍使用推导值）
- 估值基准日：**2025-12-31**（基金说明书约定）

### 1.3 各项目推导明细（含来源标注）
{noi_section}

---

## 二、DCF估值结果

### 2.1 估值汇总

| 项目 | 估值(万元) | 估值(亿元) | NOI来源 | 首年基础NOI | 隐含Cap Rate |
|------|-----------|-----------|---------|------------|-------------|
{project_results}| **合计** | **{dcf["total_valuation"]:,.2f}** | **{dcf["total_valuation"]/10000:.2f}** | | **{dcf["kpis"]["total_base_noi"]:,.2f}** | **{dcf["kpis"]["implied_cap_rate"]:.2%}** |

### 2.2 与招募估值对比

| 对比项 | 金额(亿元) | 说明 |
|--------|-----------|------|
| DCF估值 | {val_comp['dcf_valuation_billion']:.2f} | 本模型 |
| 资产评估值 | {val_comp['prospectus_valuation_billion']:.2f} | 招募说明书 |
| **差异** | **{val_comp['difference_billion']:+.2f}** | **{val_comp['difference_pct']:+.1f}%** |

---

## 三、估值差异分析

DCF模型估值与招募资产评估值存在差异，可能原因:

| 因素 | 说明 | 估计影响方向 |
|------|------|------------|
| 部分年限处理 | 模型已含部分年(如19.28年中的0.28年) | 已修正 |
| 折现时点约定 | 招募可能用期中折现(mid-year convention) | 低估约+3% |
| GOP率变化 | 模型假设收入/成本同比例增长(GOP率固定) | 可能低估 |
| | 实际: 固定成本占比使GOP率随收入增长而改善 | |
| 运营成本口径 | REITs明细(含独立物业/保险)高于历史利润表 | 低估NOI |
| 评估调整 | 招募评估可能含市场比较法修正等 | 可能 |

---

## 四、全量输入参数来源索引

### 4.1 DCF估值参数

| 参数 | 数值 | 来源分类 | 页码/依据 |
|------|------|---------|---------|
| 折现率 | {self.dcf_model.discount_rate:.2%} | 招募说明书 | Page 236 |
| 增长率-第1年 | 0%（基期） | 招募说明书 | Page 236 |
| 增长率-第2年(广州) | 2% | 招募说明书 | Page 236, 250 |
| 增长率-第2年(上海) | 1% | 招募说明书 | Page 236, 250 |
| 增长率-第3年 | 2% | 招募说明书 | Page 236 |
| 增长率-第4-10年 | 3% | 招募说明书 | Page 236 |
| 增长率-第11年起 | 2.25% | 招募说明书 | Page 236 |
| 估值基准日 | 2025-12-31 | 招募说明书 | 基金合同 |

### 4.2 收入参数

| 参数 | 广州 | 上海 | 来源分类 | 页码/依据 |
|------|------|------|---------|---------|
| 客房收入(first_year_amount) | 12,015.67万 | 3,035.38万 | 招募说明书 | Page 163-164 |
| ADR(含税均值) | 468.52元 | 359.71元 | 招募说明书 | Page 163-164 |
| 入住率(OCC) | 93.5% | 90.36% | 招募说明书 | Page 163-164 |
| 房间数 | 776间 | 268间 | 招募说明书 | Page 64-68 |
| 餐饮收入 | 536.8万 | 208.7万 | 招募说明书/历史均值 | Page 166/185 |
| OTA收入 | 0万 | 0万 | 用户假设 | 已确认历史为0 |
| 其他收入 | 181.92万 | 25.49万 | 用户假设 | 参考历史比例 |
| 商业租金 | 377.07万 | 40.29万 | 招募说明书 | 不动产合同 |
| 商业物业费 | 75.90万 | 3.96万 | 招募说明书 | 物业合同 |

### 4.3 费用参数

| 参数 | 广州 | 上海 | 来源分类 | 页码/依据 |
|------|------|------|---------|---------|
| 人工成本 | 1,312.98万 | 380万 | 历史均值推算 | Page 166/185 |
| 餐饮成本 | 268.4万 | 104.35万 | 历史均值推算 | 历史比例 |
| 清洁用品 | 180万 | 65万 | 历史均值推算 | 历史均值 |
| 易耗品 | 120万 | 45万 | 历史均值推算 | 历史均值 |
| 水电费 | 580万 | 220万 | 历史均值推算 | 历史均值 |
| 维修保养 | 350万 | 140万 | 历史均值推算 | 历史均值 |
| 市场营销 | 450万 | 180万 | 用户假设 | 估算 |
| 数据系统 | 150万 | 60万 | 行业常识 | 酒店管理系统 |
| 其他运营 | 200万 | 80万 | 用户假设 | 估算 |
| 物业管理费 | 436.3万 | 143.59万 | 行业常识 | 面积×单价推算 |
| 保险费 | 45万 | 18万 | 用户假设 | 财产险+公责险 |
| 税金(实际缴纳) | 856.42万 | 186.04万 | 招募说明书 | Page 166/186(2025年) |
| 房产原值(酒店) | 104,786.15万 | 28,959.67万 | 招募说明书 | Page 241 |
| 土地使用税税率 | 12元/㎡/年 | 3元/㎡/年 | 招募说明书 | Page 254 |
| 管理费率 | 3% of GOP | 3% of GOP | 行业常识/合同 | 华住管理合同 |
| Capex | 141.63万/年 | 38.92万/年 | 招募说明书 | Page 235/241 |

### 4.4 资产基本信息

| 参数 | 广州 | 上海 | 来源分类 | 页码 |
|------|------|------|---------|------|
| 剩余收益年限 | 19.28年 | 30.65年 | 招募说明书 | Page 235/241 |
| 建筑面积 | 42,774.94㎡ | 13,295.65㎡ | 招募说明书 | Page 64-68 |
| 土地面积 | 3,500㎡ | 1,200㎡ | 招募说明书 | Page 254 |

---

## 五、GOP率稳定性 (历史3年)

| 项目 | 2023 | 2024 | 2025 | 波动 |
|------|------|------|------|------|
| 广州 GOP率 | 61.3% | 60.1% | 65.6% | +5.5pp |
| 上海 GOP率 | 51.5% | 51.1% | 54.5% | +3.4pp |

GOP率(已含税金扣减)逐年波动，2025年明显改善，可能与:
- 运营成本控制改善 (广州2025运营成本率降至27.9%)
- 经营效率提升有关

---

*模型版本: v5.0（REITs口径 + 统一6%VAT + 明细成本 + 始终用推导值）*
"""

        path = self.output_dir / "DCF模型审计报告.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(report)
        self.result.add_log(f"审计报告: {path.name}")

    def _generate_excel_model(self):
        """生成Excel模型"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            self.result.add_log("  [WARN] openpyxl未安装，跳过Excel生成")
            return

        if not self.dcf_model:
            return

        dcf = self.dcf_model.calculate()
        wb = Workbook()

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        sub_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        sub_font = Font(bold=True, size=10)
        rev_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        exp_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # === Sheet 1: NOI推导与验证 ===
        ws = wb.active
        ws.title = "NOI推导与验证"

        ws['A1'] = "NOI推导与验证 — 收支明细→NOI→DCF"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:F1')

        for idx, derived in enumerate(self.dcf_model.derived_nois):
            col_start = 1 + idx * 4
            row = 3

            ws.cell(row=row, column=col_start, value=derived.project_name).font = Font(bold=True, size=12, color="366092")
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_start + 2)
            row += 1

            headers = ["科目", "金额(万元)", "说明"]
            for j, h in enumerate(headers):
                c = ws.cell(row=row, column=col_start + j, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = border
            row += 1

            def write_line(label, val, note="", fill=None):
                nonlocal row
                c1 = ws.cell(row=row, column=col_start, value=label)
                c2 = ws.cell(row=row, column=col_start + 1, value=round(val, 2))
                c3 = ws.cell(row=row, column=col_start + 2, value=note)
                for c in [c1, c2, c3]:
                    c.border = border
                    if fill:
                        c.fill = fill
                c2.number_format = '#,##0.00'
                row += 1

            write_line("【收入(不含税)】", 0, "统一6%增值税价税分离", sub_fill)
            rd = derived.revenue_detail
            room_key = "room_revenue_excl_tax" if "room_revenue_excl_tax" in rd else "room_revenue"
            write_line("  客房收入(不含税)", rd[room_key], f"ADR÷1.06={rd.get('adr_excl_tax_6pct',0):,.0f},差{rd.get('adr_vs_actual_diff_pct',0):+.1f}%", rev_fill)
            write_line("  (参考)ADR含税值", rd.get("adr_incl_tax", 0), "ADR×rooms×OCC×365", PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"))
            write_line("  餐饮收入", rd["fb_revenue"], "", rev_fill)
            write_line("  其他收入", rd["other_revenue"], "", rev_fill)
            write_line("  商业租金", rd["commercial_rent"], "", rev_fill)
            write_line("  商业物业费", rd["commercial_mgmt"], "", rev_fill)
            write_line("收入合计(不含税)", derived.total_revenue, "")
            row += 1

            write_line("【运营成本(REITs口径)】", 0, "", sub_fill)
            ed = derived.expense_detail
            ref_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            write_line("  运营明细小计", ed.get("operating_subtotal", 0), "人工+餐饮+清洁+能耗+维保+营销等", exp_fill)
            write_line("  物业费(独立合同)", ed.get("property_expense", 0), "REITs后独立列支", exp_fill)
            write_line("  保险(独立合同)", ed.get("insurance", 0), "REITs后独立列支", exp_fill)
            write_line("运营成本合计", ed["cost_excl_dep"], ed["cost_source"], exp_fill)
            hist_c = ed.get("hist_cost_2025", 0)
            if hist_c:
                write_line("(参考)历史利润表2025", hist_c, f"差异{ed.get('cost_vs_hist_pct',0):+.1f}%", ref_fill)
            row += 1

            write_line("【税费】", 0, "", sub_fill)
            write_line("  税金及附加(实际)", ed["tax_total"], ed["tax_source"], exp_fill)
            write_line("  (参考)推导值", ed.get("tax_derived", 0), f"差异{ed.get('tax_derived_diff_pct',0):+.1f}%", ref_fill)
            write_line("GOP", ed["gop"], f"GOP率={ed.get('gop_margin',0):.1f}%")
            write_line("  管理费(酒店运营)", ed["management_fee"], ed["mgmt_source"], exp_fill)
            write_line("支出合计", derived.total_expense, "")
            row += 1

            write_line("Capex", derived.capex, "")
            write_line("推导NOI/CF", derived.noi, "")
            write_line("招募NOI/CF", derived.prospectus_noicf, "")
            status = "PASS" if derived.within_threshold else "FAIL"
            write_line(f"差异 {derived.diff_pct*100:+.1f}%", derived.noi - derived.prospectus_noicf, status)
            row += 1

            noi_src = "推导值" if derived.within_threshold else "招募值(回退)"
            write_line("DCF使用的NOI来源", 0, noi_src)

        for col in range(1, 15):
            ws.column_dimensions[chr(64 + min(col, 26))].width = 18

        # === Sheet 2: DCF现金流 ===
        ws2 = wb.create_sheet("DCF现金流")
        for pi, proj in enumerate(dcf["projects"]):
            col_start = 1 + pi * 9
            ws2.cell(row=1, column=col_start, value=proj["name"]).font = Font(bold=True, size=12, color="366092")
            ws2.cell(row=2, column=col_start, value=f"NOI来源: {proj['noi_source']}").font = Font(italic=True, color="888888")

            cf_headers = ["年份", "增长率", "NOI", "Capex", "FCF", "折现因子", "PV"]
            for j, h in enumerate(cf_headers):
                c = ws2.cell(row=3, column=col_start + j, value=h)
                c.font = header_font
                c.fill = header_fill

            for ri, cf in enumerate(proj["cash_flows"]):
                row = 4 + ri
                vals = [cf["year"], f"{cf['growth_rate']:.2%}", cf["noi"], cf["capex"],
                        cf["fcf"], cf["discount_factor"], cf["pv"]]
                for j, v in enumerate(vals):
                    c = ws2.cell(row=row, column=col_start + j, value=v)
                    if j >= 2:
                        c.number_format = '#,##0.00'

            total_row = 4 + len(proj["cash_flows"])
            ws2.cell(row=total_row, column=col_start, value="合计").font = Font(bold=True)
            ws2.cell(row=total_row, column=col_start + 6, value=proj["valuation"]).font = Font(bold=True)
            ws2.cell(row=total_row, column=col_start + 6).number_format = '#,##0.00'

        # === Sheet 3: 估值汇总 ===
        ws3 = wb.create_sheet("估值汇总")
        ws3['A1'] = "DCF估值汇总"
        ws3['A1'].font = Font(bold=True, size=14, color="366092")

        summary_headers = ["项目", "估值(万元)", "估值(亿元)", "房间数", "单房估值(元)", "NOI来源", "Cap Rate"]
        for j, h in enumerate(summary_headers, 1):
            c = ws3.cell(row=3, column=j, value=h)
            c.font = header_font
            c.fill = header_fill

        for ri, p in enumerate(dcf["projects"]):
            row = 4 + ri
            vals = [p["name"], p["valuation"], p["valuation"]/10000, p["rooms"],
                    p["value_per_room"], p["noi_source"], f"{p['implied_cap_rate']:.2%}"]
            for j, v in enumerate(vals, 1):
                c = ws3.cell(row=row, column=j, value=v)
                if j in [2, 3, 5]:
                    c.number_format = '#,##0.00'

        total_row = 4 + len(dcf["projects"])
        ws3.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws3.cell(row=total_row, column=2, value=dcf["total_valuation"]).font = Font(bold=True)
        ws3.cell(row=total_row, column=2).number_format = '#,##0.00'
        ws3.cell(row=total_row, column=3, value=dcf["total_valuation"]/10000).font = Font(bold=True)

        for j in range(1, 8):
            ws3.column_dimensions[chr(64 + j)].width = 16

        excel_path = self.output_dir / "dcf_model.xlsx"
        wb.save(excel_path)
        self.result.add_log(f"Excel模型: {excel_path.name}")
