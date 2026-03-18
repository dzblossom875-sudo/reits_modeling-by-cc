"""
Excel文件生成器（方式A）
使用openpyxl直接生成投行风格的Excel模型
"""

from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..core.types import ValuationResult, ScenarioResult
from ..core.config import EXCEL_TEMPLATE_CONFIG
from ..core.exceptions import ExportError
from .data_provider import DataProvider


class ExcelGenerator:
    """Excel文件生成器"""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ExportError("openpyxl未安装，请运行: pip install openpyxl")

        self.wb = Workbook()
        self.data_provider = DataProvider()

        # 定义样式
        self._setup_styles()

    def _setup_styles(self):
        """设置单元格样式"""
        # 标题样式
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, size=12, color="FFFFFF")
        self.header_style.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")

        # 子标题样式
        self.subheader_style = NamedStyle(name="subheader")
        self.subheader_style.font = Font(bold=True, size=11)
        self.subheader_style.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # 数字样式
        self.number_style = NamedStyle(name="number")
        self.number_style.number_format = '#,##0.00'

        # 百分比样式
        self.percent_style = NamedStyle(name="percent")
        self.percent_style.number_format = '0.00%'

        # 货币样式
        self.currency_style = NamedStyle(name="currency")
        self.currency_style.number_format = '#,##0.00'

        # 边框样式
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(
        self,
        valuation: ValuationResult,
        scenarios: Optional[List[ScenarioResult]] = None,
        output_path: str = "reit_valuation.xlsx"
    ) -> str:
        """
        生成Excel估值模型

        Args:
            valuation: 基准情景估值结果
            scenarios: 其他情景结果（可选）
            output_path: 输出文件路径

        Returns:
            生成的文件路径
        """
        # 删除默认sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # 创建各sheet
        self._create_dashboard(valuation)
        self._create_assumptions(valuation)
        self._create_dcf(valuation)

        if scenarios:
            self._create_scenarios(scenarios)

        self._create_data(valuation)

        # 保存文件
        self.wb.save(output_path)
        return output_path

    def _create_dashboard(self, valuation: ValuationResult):
        """创建Dashboard Sheet"""
        ws = self.wb.create_sheet("Dashboard")

        # 标题
        ws['A1'] = "REITs估值模型"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # 项目信息
        ws['A3'] = "项目名称"
        ws['B3'] = valuation.project_info.name or "未命名项目"
        ws['A4'] = "资产类型"
        ws['B4'] = valuation.asset_type.value
        ws['A5'] = "估值日期"
        ws['B5'] = valuation.created_at.strftime("%Y-%m-%d")

        # 关键指标
        ws['A7'] = "关键估值指标"
        ws['A7'].style = self.subheader_style
        ws.merge_cells('A7:B7')

        ws['A8'] = "DCF估值（万元）"
        ws['B8'] = round(valuation.dcf_value, 2)
        ws['B8'].style = self.currency_style

        ws['A9'] = "NPV（万元）"
        ws['B9'] = round(valuation.npv, 2)
        ws['B9'].style = self.currency_style

        if valuation.irr:
            ws['A10'] = "IRR"
            ws['B10'] = valuation.irr
            ws['B10'].style = self.percent_style

        if valuation.cap_rate:
            ws['A11'] = "资本化率估值"
            ws['B11'] = valuation.cap_rate
            ws['B11'].style = self.percent_style

        # 现金流汇总
        ws['A13'] = "现金流汇总"
        ws['A13'].style = self.subheader_style
        ws.merge_cells('A13:B13')

        total_noi = sum(cf.calculate_noi() for cf in valuation.cash_flows)
        avg_noi = total_noi / len(valuation.cash_flows) if valuation.cash_flows else 0

        ws['A14'] = "10年总NOI（万元）"
        ws['B14'] = round(total_noi, 2)
        ws['B14'].style = self.currency_style

        ws['A15'] = "平均年NOI（万元）"
        ws['B15'] = round(avg_noi, 2)
        ws['B15'].style = self.currency_style

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_assumptions(self, valuation: ValuationResult):
        """创建Assumptions Sheet"""
        ws = self.wb.create_sheet("Assumptions")

        # 标题
        ws['A1'] = "关键假设"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:C1')

        # 表头
        headers = ["参数", "数值", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.style = self.subheader_style

        # 数据
        row = 4
        for key, value in valuation.assumptions.items():
            ws.cell(row=row, column=1, value=key)

            # 根据值的类型设置格式
            cell = ws.cell(row=row, column=2)
            if isinstance(value, str) and '%' in value:
                # 百分比
                try:
                    pct_value = float(value.replace('%', '')) / 100
                    cell.value = pct_value
                    cell.style = self.percent_style
                except:
                    cell.value = value
            elif isinstance(value, (int, float)):
                cell.value = value
                cell.style = self.number_style
            else:
                cell.value = str(value)

            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30

    def _create_dcf(self, valuation: ValuationResult):
        """创建DCF Sheet"""
        ws = self.wb.create_sheet("DCF")

        # 标题
        ws['A1'] = "DCF现金流折现计算"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:K1')

        # 表头
        headers = [
            "年份", "租金收入", "其他收入", "总收入",
            "运营费用", "管理费用", "NOI", "折现因子", "现值"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.style = self.subheader_style

        # 现金流数据
        discount_rate = 0.075  # 假设折现率7.5%

        for idx, cf in enumerate(valuation.cash_flows, start=4):
            row = idx

            ws.cell(row=row, column=1, value=cf.year)
            ws.cell(row=row, column=2, value=round(cf.rental_income, 2)).style = self.number_style
            ws.cell(row=row, column=3, value=round(cf.other_income, 2)).style = self.number_style
            ws.cell(row=row, column=4, value=round(cf.total_income, 2)).style = self.number_style
            ws.cell(row=row, column=5, value=round(cf.operating_expense, 2)).style = self.number_style
            ws.cell(row=row, column=6, value=round(cf.management_fee, 2)).style = self.number_style

            noi = cf.calculate_noi()
            ws.cell(row=row, column=7, value=round(noi, 2)).style = self.number_style

            # 折现因子公式
            discount_factor = 1 / ((1 + discount_rate) ** cf.year)
            ws.cell(row=row, column=8, value=round(discount_factor, 6)).style = self.number_style

            # 现值
            pv = noi * discount_factor
            ws.cell(row=row, column=9, value=round(pv, 2)).style = self.number_style

        # 终值计算
        terminal_row = 4 + len(valuation.cash_flows)
        if valuation.cash_flows:
            final_noi = valuation.cash_flows[-1].calculate_noi()
            terminal_growth = 0.02
            terminal_value = final_noi * (1 + terminal_growth) / (discount_rate - terminal_growth)
            terminal_pv = terminal_value / ((1 + discount_rate) ** len(valuation.cash_flows))

            ws.cell(row=terminal_row, column=1, value="终值")
            ws.cell(row=terminal_row, column=7, value=round(terminal_value, 2)).style = self.number_style
            ws.cell(row=terminal_row, column=9, value=round(terminal_pv, 2)).style = self.number_style

        # NPV总计
        npv_row = terminal_row + 2
        ws.cell(row=npv_row, column=1, value="NPV")
        ws.cell(row=npv_row, column=1).font = Font(bold=True)
        ws.cell(row=npv_row, column=9, value=round(valuation.npv, 2)).style = self.currency_style
        ws.cell(row=npv_row, column=9).font = Font(bold=True)

        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_scenarios(self, scenarios: List[ScenarioResult]):
        """创建Scenarios Sheet"""
        ws = self.wb.create_sheet("Scenarios")

        # 标题
        ws['A1'] = "情景对比分析"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:E1')

        # 表头
        headers = ["情景", "NPV（万元）", "IRR", "vs基准", "主要假设调整"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.style = self.subheader_style

        # 数据
        for idx, scenario in enumerate(scenarios, start=4):
            row = idx

            ws.cell(row=row, column=1, value=scenario.scenario_name)
            ws.cell(row=row, column=2, value=round(scenario.valuation.npv, 2)).style = self.currency_style

            if scenario.valuation.irr:
                ws.cell(row=row, column=3, value=scenario.valuation.irr).style = self.percent_style

            if scenario.vs_base_percent is not None:
                vs_base = scenario.vs_base_percent
                cell = ws.cell(row=row, column=4, value=vs_base)
                cell.style = self.percent_style
                cell.number_format = '+0.00%;-0.00%;0.00%'

            # 假设调整摘要
            adjustments_str = ", ".join([
                f"{k}={v}" for k, v in scenario.valuation.assumptions.items()
            ][:3])  # 只显示前3个
            ws.cell(row=row, column=5, value=adjustments_str)

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40

    def _create_data(self, valuation: ValuationResult):
        """创建Data Sheet（数据来源记录）"""
        ws = self.wb.create_sheet("Data")

        # 标题
        ws['A1'] = "数据来源与假设记录"
        ws['A1'].style = self.header_style
        ws.merge_cells('A1:C1')

        # 时间戳
        ws['A3'] = "生成时间"
        ws['B3'] = valuation.created_at.strftime("%Y-%m-%d %H:%M:%S")

        # 计算说明
        if valuation.calculation_notes:
            ws['A5'] = "计算说明"
            ws['A5'].style = self.subheader_style

            for idx, note in enumerate(valuation.calculation_notes, start=6):
                ws.cell(row=idx, column=1, value=note)

        # 原始假设
        ws['A15'] = "原始假设记录"
        ws['A15'].style = self.subheader_style

        row = 16
        for key, value in valuation.assumptions.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
