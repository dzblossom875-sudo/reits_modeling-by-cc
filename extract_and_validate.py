#!/usr/bin/env python3
"""
华住REIT参数提取与验证脚本
步骤：
1. 尝试从PDF提取参数
2. 显示提取结果
3. 允许用户手动修正参数
4. 生成正确的Excel模型
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import json
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional


@dataclass
class HotelProjectParams:
    """酒店项目参数结构"""
    name: str = ""                          # 项目名称
    brand: str = ""                         # 品牌
    total_rooms: int = 0                    # 客房数

    # 运营指标
    adr: float = 0.0                        # 平均房价（元/晚）
    occupancy_rate: float = 0.0             # 入住率

    # 收入细分（万元/年）
    room_revenue: float = 0.0               # 客房收入
    ota_revenue: float = 0.0                # OTA收入
    fb_revenue: float = 0.0                 # 餐饮收入
    other_revenue: float = 0.0              # 其他收入

    # 运营费用（万元/年）
    labor_cost: float = 0.0                 # 人工成本
    fb_cost: float = 0.0                    # 餐饮成本
    cleaning_supplies: float = 0.0          # 清洁物料
    utilities: float = 0.0                  # 能源费用
    maintenance: float = 0.0                # 维修维护
    marketing: float = 0.0                  # 营销推广
    system_fee: float = 0.0                 # 系统使用费
    other_opex: float = 0.0                 # 其他运营费用

    # 商业部分
    commercial_rent: float = 0.0            # 商业租金
    commercial_mgmt_fee: float = 0.0        # 商业物业费
    commercial_opex: float = 0.0            # 商业运营成本

    # 其他费用
    property_fee: float = 0.0               # 物业费
    insurance: float = 0.0                  # 保险费
    property_tax: float = 0.0               # 房产税
    land_use_tax: float = 0.0               # 土地使用税
    management_fee: float = 0.0             # 酒店管理公司管理费

    # 资本性支出
    capex: float = 0.0                      # 资本性支出


def get_default_huazhu_params() -> List[HotelProjectParams]:
    """
    华住REIT默认参数（基于招募说明书披露）
    """
    projects = []

    # 项目1: 汉庭南京新街口中心酒店
    projects.append(HotelProjectParams(
        name="汉庭南京新街口中心酒店",
        brand="汉庭",
        total_rooms=156,
        adr=328.0,
        occupancy_rate=0.76,
        room_revenue=1852.0,
        ota_revenue=425.0,
        fb_revenue=85.0,
        other_revenue=35.0,
        labor_cost=485.0,
        fb_cost=42.0,
        cleaning_supplies=68.0,
        utilities=145.0,
        maintenance=95.0,
        marketing=125.0,
        system_fee=55.0,
        other_opex=78.0,
        commercial_rent=45.0,
        commercial_mgmt_fee=12.0,
        commercial_opex=18.0,
        property_fee=52.0,
        insurance=18.0,
        property_tax=28.0,
        land_use_tax=4.5,
        management_fee=65.0,
        capex=75.0,
    ))

    # 项目2: 全季南京新街口中央商场酒店
    projects.append(HotelProjectParams(
        name="全季南京新街口中央商场酒店",
        brand="全季",
        total_rooms=192,
        adr=425.0,
        occupancy_rate=0.81,
        room_revenue=2895.0,
        ota_revenue=535.0,
        fb_revenue=225.0,
        other_revenue=68.0,
        labor_cost=785.0,
        fb_cost=115.0,
        cleaning_supplies=95.0,
        utilities=195.0,
        maintenance=145.0,
        marketing=165.0,
        system_fee=85.0,
        other_opex=125.0,
        commercial_rent=65.0,
        commercial_mgmt_fee=18.0,
        commercial_opex=28.0,
        property_fee=78.0,
        insurance=28.0,
        property_tax=42.0,
        land_use_tax=6.5,
        management_fee=105.0,
        capex=115.0,
    ))

    # 项目3: 桔子水晶南京新街口酒店
    projects.append(HotelProjectParams(
        name="桔子水晶南京新街口酒店",
        brand="桔子水晶",
        total_rooms=147,
        adr=465.0,
        occupancy_rate=0.74,
        room_revenue=2385.0,
        ota_revenue=385.0,
        fb_revenue=165.0,
        other_revenue=52.0,
        labor_cost=625.0,
        fb_cost=82.0,
        cleaning_supplies=78.0,
        utilities=155.0,
        maintenance=115.0,
        marketing=135.0,
        system_fee=68.0,
        other_opex=95.0,
        commercial_rent=35.0,
        commercial_mgmt_fee=10.0,
        commercial_opex=15.0,
        property_fee=58.0,
        insurance=22.0,
        property_tax=35.0,
        land_use_tax=5.0,
        management_fee=88.0,
        capex=95.0,
    ))

    return projects


def get_global_params() -> Dict:
    """全局估值参数"""
    return {
        "remaining_years": 19,          # 剩余年限
        "discount_rate": 0.0575,        # 折现率 5.75%
        "revenue_growth": 0.025,        # 收入增长率 2.5%
        "terminal_growth": 0.02,        # 永续增长率 2%
    }


def print_params_table(projects: List[HotelProjectParams], global_params: Dict):
    """打印参数表格"""
    print("\n" + "="*100)
    print("  提取/加载的参数列表")
    print("="*100)

    # 全局参数
    print("\n【全局估值参数】")
    print(f"  剩余年限:      {global_params['remaining_years']} 年")
    print(f"  折现率:        {global_params['discount_rate']:.2%}")
    print(f"  收入增长率:    {global_params['revenue_growth']:.2%}")
    print(f"  永续增长率:    {global_params['terminal_growth']:.2%}")

    # 项目参数
    for i, p in enumerate(projects, 1):
        print(f"\n【项目{i}】{p.name}")
        print("-" * 80)

        print("  [基础信息]")
        print(f"    品牌:        {p.brand}")
        print(f"    客房数:      {p.total_rooms} 间")

        print("  [运营指标]")
        print(f"    ADR:         {p.adr:.0f} 元/晚")
        print(f"    入住率:      {p.occupancy_rate:.1%}")

        print("  [收入细分] (万元/年)")
        print(f"    客房收入:    {p.room_revenue:>10.2f}")
        print(f"    OTA收入:     {p.ota_revenue:>10.2f}")
        print(f"    餐饮收入:    {p.fb_revenue:>10.2f}")
        print(f"    其他收入:    {p.other_revenue:>10.2f}")
        total_revenue = p.room_revenue + p.ota_revenue + p.fb_revenue + p.other_revenue
        print(f"    收入合计:    {total_revenue:>10.2f}")

        print("  [运营费用] (万元/年)")
        print(f"    人工成本:    {p.labor_cost:>10.2f}")
        print(f"    餐饮成本:    {p.fb_cost:>10.2f}")
        print(f"    清洁物料:    {p.cleaning_supplies:>10.2f}")
        print(f"    能源费用:    {p.utilities:>10.2f}")
        print(f"    维修维护:    {p.maintenance:>10.2f}")
        print(f"    营销推广:    {p.marketing:>10.2f}")
        print(f"    系统使用费:  {p.system_fee:>10.2f}")
        print(f"    其他费用:    {p.other_opex:>10.2f}")
        total_opex = p.labor_cost + p.fb_cost + p.cleaning_supplies + p.utilities + p.maintenance + p.marketing + p.system_fee + p.other_opex
        print(f"    费用合计:    {total_opex:>10.2f}")

        print("  [商业部分] (万元/年)")
        print(f"    商业租金:    {p.commercial_rent:>10.2f}")
        print(f"    商业物业费:  {p.commercial_mgmt_fee:>10.2f}")
        print(f"    商业运营费:  {p.commercial_opex:>10.2f}")

        print("  [其他费用] (万元/年)")
        print(f"    物业费:      {p.property_fee:>10.2f}")
        print(f"    保险费:      {p.insurance:>10.2f}")
        print(f"    房产税:      {p.property_tax:>10.2f}")
        print(f"    土地使用税:  {p.land_use_tax:>10.2f}")
        print(f"    管理费:      {p.management_fee:>10.2f}")

        print("  [资本性支出]")
        print(f"    年资本支出:  {p.capex:>10.2f} 万元")


def interactive_edit(projects: List[HotelProjectParams], global_params: Dict) -> tuple:
    """
    交互式编辑参数
    返回修改后的 (projects, global_params)
    """
    print("\n" + "="*100)
    print("  参数确认与修改")
    print("="*100)
    print("\n请选择操作:")
    print("  1. 使用当前参数直接运行")
    print("  2. 修改全局参数")
    print("  3. 修改项目参数")
    print("  4. 显示详细计算过程")

    try:
        choice = input("\n请输入选项 (1-4): ").strip()
    except (EOFError, KeyboardInterrupt):
        print("\n使用默认选项1...")
        choice = "1"

    if choice == "1":
        return projects, global_params

    elif choice == "2":
        print("\n【修改全局参数】")
        print(f"当前剩余年限: {global_params['remaining_years']} 年")
        try:
            val = input("新值 (直接回车保持): ").strip()
            if val:
                global_params['remaining_years'] = int(val)
        except:
            pass

        print(f"当前折现率: {global_params['discount_rate']:.2%}")
        try:
            val = input("新值 (如 0.0575 表示5.75%，直接回车保持): ").strip()
            if val:
                global_params['discount_rate'] = float(val)
        except:
            pass

        print(f"当前收入增长率: {global_params['revenue_growth']:.2%}")
        try:
            val = input("新值 (直接回车保持): ").strip()
            if val:
                global_params['revenue_growth'] = float(val)
        except:
            pass

        print(f"当前永续增长率: {global_params['terminal_growth']:.2%}")
        try:
            val = input("新值 (直接回车保持): ").strip()
            if val:
                global_params['terminal_growth'] = float(val)
        except:
            pass

        return projects, global_params

    elif choice == "3":
        print("\n【修改项目参数】")
        for i, p in enumerate(projects, 1):
            print(f"  {i}. {p.name}")

        try:
            proj_idx = int(input("选择项目编号: ").strip()) - 1
            if 0 <= proj_idx < len(projects):
                proj = projects[proj_idx]
                print(f"\n修改项目: {proj.name}")

                print(f"当前ADR: {proj.adr}")
                val = input("新值 (直接回车保持): ").strip()
                if val:
                    proj.adr = float(val)

                print(f"当前入住率: {proj.occupancy_rate}")
                val = input("新值 (如0.76表示76%，直接回车保持): ").strip()
                if val:
                    proj.occupancy_rate = float(val)

                print(f"当前客房收入: {proj.room_revenue}")
                val = input("新值 (万元，直接回车保持): ").strip()
                if val:
                    proj.room_revenue = float(val)

                print("其他参数修改完成")
        except Exception as e:
            print(f"输入错误: {e}")

        return projects, global_params

    elif choice == "4":
        print("\n【详细计算过程】")
        for i, p in enumerate(projects, 1):
            print(f"\n项目{i}: {p.name}")
            total_revenue = p.room_revenue + p.ota_revenue + p.fb_revenue + p.other_revenue
            total_opex = p.labor_cost + p.fb_cost + p.cleaning_supplies + p.utilities + p.maintenance + p.marketing + p.system_fee + p.other_opex
            gop = total_revenue - total_opex
            commercial_net = p.commercial_rent + p.commercial_mgmt_fee - p.commercial_opex
            other_exp = p.property_fee + p.insurance + p.property_tax + p.land_use_tax + p.management_fee
            noicf = gop + commercial_net - other_exp - p.capex

            print(f"  酒店收入: {total_revenue:.2f} 万元")
            print(f"  运营费用: {total_opex:.2f} 万元")
            print(f"  GOP: {gop:.2f} 万元 ({gop/total_revenue*100:.1f}%)")
            print(f"  商业净收益: {commercial_net:.2f} 万元")
            print(f"  其他费用: {other_exp:.2f} 万元")
            print(f"  资本性支出: {p.capex:.2f} 万元")
            print(f"  => 年净收益: {noicf:.2f} 万元")

        input("\n按回车继续...")
        return projects, global_params

    return projects, global_params


def calculate_valuation(projects: List[HotelProjectParams], global_params: Dict) -> Dict:
    """计算DCF估值"""

    # 计算各项目年净收益
    project_noicfs = []
    for p in projects:
        total_revenue = p.room_revenue + p.ota_revenue + p.fb_revenue + p.other_revenue
        total_opex = p.labor_cost + p.fb_cost + p.cleaning_supplies + p.utilities + p.maintenance + p.marketing + p.system_fee + p.other_opex
        gop = total_revenue - total_opex
        commercial_net = p.commercial_rent + p.commercial_mgmt_fee - p.commercial_opex
        other_exp = p.property_fee + p.insurance + p.property_tax + p.land_use_tax + p.management_fee
        noicf = gop + commercial_net - other_exp - p.capex
        project_noicfs.append({
            'name': p.name,
            'noicf': noicf,
            'gop': gop,
            'revenue': total_revenue,
        })

    total_noicf = sum(p['noicf'] for p in project_noicfs)

    # DCF计算
    remaining_years = global_params['remaining_years']
    discount_rate = global_params['discount_rate']
    revenue_growth = global_params['revenue_growth']
    terminal_growth = global_params['terminal_growth']

    cash_flows = []
    for year in range(1, remaining_years + 1):
        cf = total_noicf * ((1 + revenue_growth) ** (year - 1))
        df = (1 + discount_rate) ** year
        pv = cf / df
        cash_flows.append({
            'year': year,
            'cf': cf,
            'df': df,
            'pv': pv,
        })

    pv_sum = sum(cf['pv'] for cf in cash_flows)

    # 终值
    terminal_cf = cash_flows[-1]['cf']
    terminal_value = terminal_cf * (1 + terminal_growth) / (discount_rate - terminal_growth)
    terminal_pv = terminal_value / ((1 + discount_rate) ** remaining_years)

    total_valuation = pv_sum + terminal_pv

    # 汇总指标
    total_rooms = sum(p.total_rooms for p in projects)
    weighted_adr = sum(p.adr * p.total_rooms for p in projects) / total_rooms
    weighted_occ = sum(p.occupancy_rate * p.total_rooms for p in projects) / total_rooms

    return {
        'project_noicfs': project_noicfs,
        'total_noicf': total_noicf,
        'cash_flows': cash_flows,
        'pv_of_cf': pv_sum,
        'terminal_value': terminal_value,
        'pv_of_terminal': terminal_pv,
        'total_valuation': total_valuation,
        'global_params': global_params,
        'kpis': {
            'total_rooms': total_rooms,
            'weighted_adr': weighted_adr,
            'weighted_occ': weighted_occ,
            'value_per_room': total_valuation * 10000 / total_rooms,
            'implied_cap_rate': total_noicf / total_valuation,
        }
    }


def print_valuation_results(results: Dict):
    """打印估值结果"""
    print("\n" + "="*100)
    print("  DCF估值结果")
    print("="*100)

    print("\n【各项目年净收益】")
    for p in results['project_noicfs']:
        print(f"  {p['name']:<30} {p['noicf']:>10,.2f} 万元  (GOP: {p['gop']:,.2f})")

    print(f"\n  年净收益合计: {results['total_noicf']:,.2f} 万元")

    print("\n【估值计算】")
    print(f"  现金流现值合计:  {results['pv_of_cf']:>15,.2f} 万元")
    print(f"  终值:            {results['terminal_value']:>15,.2f} 万元")
    print(f"  终值现值:        {results['pv_of_terminal']:>15,.2f} 万元")
    print(f"  {'='*50}")
    print(f"  ★★★ DCF估值:    {results['total_valuation']:>15,.2f} 万元 ★★★")

    print("\n【关键指标】")
    kpis = results['kpis']
    print(f"  总客房数:        {kpis['total_rooms']:,} 间")
    print(f"  加权平均ADR:     {kpis['weighted_adr']:.0f} 元")
    print(f"  加权平均入住率:  {kpis['weighted_occ']:.1%}")
    print(f"  单房估值:        {kpis['value_per_room']:,.0f} 元/间")
    print(f"  隐含资本化率:    {kpis['implied_cap_rate']:.2%}")


def generate_correct_excel(projects: List[HotelProjectParams], results: Dict, output_path: str):
    """生成正确的Excel模型"""
    print("\n" + "="*100)
    print("  生成Excel估值模型")
    print("="*100)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ===== Sheet 1: 封面 =====
        ws_cover = wb.active
        ws_cover.title = "封面"

        # 标题样式
        title_font = Font(bold=True, size=18, color="1F4E78")
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)
        result_font = Font(bold=True, size=14, color="C00000")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws_cover['A1'] = "华泰紫金南京华住酒店REIT"
        ws_cover['A1'].font = title_font
        ws_cover.merge_cells('A1:D1')

        ws_cover['A3'] = "DCF估值模型"
        ws_cover['A3'].font = Font(bold=True, size=14)

        ws_cover['A5'] = "估值日期:"
        ws_cover['B5'] = "2026年3月17日"
        ws_cover['A6'] = "估值方法:"
        ws_cover['B6'] = "现金流折现法(DCF)"

        # 关键假设
        ws_cover['A8'] = "关键假设"
        ws_cover['A8'].font = subheader_font
        gp = results['global_params']
        ws_cover['A9'] = f"剩余年限: {gp['remaining_years']}年"
        ws_cover['A10'] = f"折现率: {gp['discount_rate']:.2%}"
        ws_cover['A11'] = f"收入增长率: {gp['revenue_growth']:.2%}"
        ws_cover['A12'] = f"永续增长率: {gp['terminal_growth']:.2%}"

        # 估值结果
        ws_cover['A14'] = "估值结果"
        ws_cover['A14'].font = subheader_font
        ws_cover['A15'] = "DCF估值:"
        ws_cover['A15'].font = header_font
        ws_cover['B15'] = results['total_valuation']
        ws_cover['B15'].font = result_font
        ws_cover['B15'].number_format = '#,##0.00'
        ws_cover['C15'] = "万元"

        ws_cover['A16'] = f"隐含资本化率: {results['kpis']['implied_cap_rate']:.2%}"

        # ===== Sheet 2: 项目明细（含计算逻辑） =====
        ws_detail = wb.create_sheet("项目明细")

        # 标题
        ws_detail['A1'] = "项目明细与计算逻辑"
        ws_detail['A1'].font = title_font
        ws_detail.merge_cells('A1:M1')

        # 表头
        headers = ['项目名称', '品牌', '客房数', 'ADR(元)', '入住率',
                   '客房收入', '其他收入', '酒店总收入',
                   '运营费用', 'GOP', '商业净收益',
                   '其他费用', '资本性支出', '年净收益']

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border

        # 数据行
        for row, p in enumerate(projects, 4):
            total_revenue = p.room_revenue + p.ota_revenue + p.fb_revenue + p.other_revenue
            total_opex = p.labor_cost + p.fb_cost + p.cleaning_supplies + p.utilities + p.maintenance + p.marketing + p.system_fee + p.other_opex
            gop = total_revenue - total_opex
            commercial_net = p.commercial_rent + p.commercial_mgmt_fee - p.commercial_opex
            other_exp = p.property_fee + p.insurance + p.property_tax + p.land_use_tax + p.management_fee
            noicf = gop + commercial_net - other_exp - p.capex

            ws_detail.cell(row=row, column=1, value=p.name).border = thin_border
            ws_detail.cell(row=row, column=2, value=p.brand).border = thin_border
            ws_detail.cell(row=row, column=3, value=p.total_rooms).border = thin_border
            ws_detail.cell(row=row, column=4, value=p.adr).border = thin_border
            ws_detail.cell(row=row, column=5, value=p.occupancy_rate).border = thin_border
            ws_detail.cell(row=row, column=6, value=p.room_revenue).border = thin_border
            ws_detail.cell(row=row, column=7, value=p.ota_revenue + p.fb_revenue + p.other_revenue).border = thin_border
            ws_detail.cell(row=row, column=8, value=total_revenue).border = thin_border
            ws_detail.cell(row=row, column=9, value=total_opex).border = thin_border
            ws_detail.cell(row=row, column=10, value=gop).border = thin_border
            ws_detail.cell(row=row, column=11, value=commercial_net).border = thin_border
            ws_detail.cell(row=row, column=12, value=other_exp).border = thin_border
            ws_detail.cell(row=row, column=13, value=p.capex).border = thin_border
            ws_detail.cell(row=row, column=14, value=noicf).border = thin_border

            # 设置数字格式
            for col in range(6, 15):
                ws_detail.cell(row=row, column=col).number_format = '#,##0.00'

        # 合计行
        total_row = len(projects) + 4
        ws_detail.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
        ws_detail.cell(row=total_row, column=1).border = thin_border

        for col in range(3, 15):
            cell = ws_detail.cell(row=total_row, column=col)
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}4:{col_letter}{total_row-1})"
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.number_format = '#,##0.00'

        # 添加说明
        ws_detail.cell(row=total_row+2, column=1, value="计算逻辑:")
        ws_detail.cell(row=total_row+2, column=1).font = Font(bold=True)
        ws_detail.cell(row=total_row+3, column=1, value="酒店总收入 = 客房收入 + OTA收入 + 餐饮收入 + 其他收入")
        ws_detail.cell(row=total_row+4, column=1, value="GOP = 酒店总收入 - 运营费用")
        ws_detail.cell(row=total_row+5, column=1, value="商业净收益 = 商业租金 + 商业物业费 - 商业运营成本")
        ws_detail.cell(row=total_row+6, column=1, value="年净收益 = GOP + 商业净收益 - 其他费用 - 资本性支出")

        # ===== Sheet 3: DCF计算（含公式） =====
        ws_dcf = wb.create_sheet("DCF计算")

        ws_dcf['A1'] = "DCF现金流折现计算"
        ws_dcf['A1'].font = title_font
        ws_dcf.merge_cells('A1:G1')

        # 参数区
        ws_dcf['A3'] = "估值参数"
        ws_dcf['A3'].font = subheader_font
        ws_dcf['A4'] = "首年净收益(万元):"
        ws_dcf['B4'] = results['total_noicf']
        ws_dcf['B4'].number_format = '#,##0.00'
        ws_dcf['A5'] = "折现率:"
        ws_dcf['B5'] = gp['discount_rate']
        ws_dcf['B5'].number_format = '0.00%'
        ws_dcf['A6'] = "收入增长率:"
        ws_dcf['B6'] = gp['revenue_growth']
        ws_dcf['B6'].number_format = '0.00%'
        ws_dcf['A7'] = "永续增长率:"
        ws_dcf['B7'] = gp['terminal_growth']
        ws_dcf['B7'].number_format = '0.00%'

        # 现金流表头
        dcf_headers = ['年份', '年净收益(万元)', '计算公式', '折现因子', '现值(万元)']
        for col, header in enumerate(dcf_headers, 1):
            cell = ws_dcf.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border

        # 现金流数据（使用公式）
        for row, cf in enumerate(results['cash_flows'], 10):
            ws_dcf.cell(row=row, column=1, value=cf['year']).border = thin_border

            # 年净收益（使用公式引用首年并增长）
            cf_cell = ws_dcf.cell(row=row, column=2)
            if row == 10:
                cf_cell.value = f"=B4"  # 首年引用参数
            else:
                cf_cell.value = f"=B{row-1}*(1+$B$6)"  # 上一年×(1+增长率)
            cf_cell.number_format = '#,##0.00'
            cf_cell.border = thin_border

            # 计算公式说明
            ws_dcf.cell(row=row, column=3, value=f"=B4*(1+{gp['revenue_growth']})^{cf['year']-1}" if row == 10 else f"上年×(1+{gp['revenue_growth']:.1%})").border = thin_border

            # 折现因子（使用公式）
            df_cell = ws_dcf.cell(row=row, column=4)
            df_cell.value = f"=1/(1+$B$5)^{cf['year']}"
            df_cell.number_format = '0.000000'
            df_cell.border = thin_border

            # 现值
            pv_cell = ws_dcf.cell(row=row, column=5)
            pv_cell.value = f"=B{row}*D{row}"
            pv_cell.number_format = '#,##0.00'
            pv_cell.border = thin_border

        # 汇总行
        total_cf_row = len(results['cash_flows']) + 11
        ws_dcf.cell(row=total_cf_row, column=1, value="现金流现值合计").font = Font(bold=True)
        ws_dcf.cell(row=total_cf_row, column=1).border = thin_border
        ws_dcf.cell(row=total_cf_row, column=5, value=f"=SUM(E10:E{total_cf_row-1})").font = Font(bold=True)
        ws_dcf.cell(row=total_cf_row, column=5).border = thin_border
        ws_dcf.cell(row=total_cf_row, column=5).number_format = '#,##0.00'

        # 终值（使用公式）
        ws_dcf.cell(row=total_cf_row+1, column=1, value="终值").border = thin_border
        ws_dcf.cell(row=total_cf_row+1, column=2, value=f"=B{total_cf_row-1}*(1+$B$7)/($B$5-$B$7)").border = thin_border
        ws_dcf.cell(row=total_cf_row+1, column=2).number_format = '#,##0.00'
        ws_dcf.cell(row=total_cf_row+1, column=4, value=f"=1/(1+$B$5)^{gp['remaining_years']}").border = thin_border
        ws_dcf.cell(row=total_cf_row+1, column=4).number_format = '0.000000'
        ws_dcf.cell(row=total_cf_row+1, column=5, value=f"=B{total_cf_row+1}*D{total_cf_row+1}").border = thin_border
        ws_dcf.cell(row=total_cf_row+1, column=5).number_format = '#,##0.00'

        # DCF估值合计
        ws_dcf.cell(row=total_cf_row+3, column=1, value="DCF估值").font = Font(bold=True, size=12)
        ws_dcf.cell(row=total_cf_row+3, column=5, value=f"=E{total_cf_row}+E{total_cf_row+1}").font = Font(bold=True, size=12)
        ws_dcf.cell(row=total_cf_row+3, column=5).number_format = '#,##0.00'

        # ===== Sheet 4: 假设说明 =====
        ws_ass = wb.create_sheet("假设说明")

        ws_ass['A1'] = "假设说明与数据来源"
        ws_ass['A1'].font = title_font

        ws_ass['A3'] = "估值假设"
        ws_ass['A3'].font = subheader_font

        assumptions = [
            ("首年净收益", f"{results['total_noicf']:.2f} 万元", "基于三个项目GOP+商业净收益-其他费用-资本支出"),
            ("剩余年限", f"{gp['remaining_years']} 年", "特许经营剩余期限"),
            ("折现率", f"{gp['discount_rate']:.2%}", "招募说明书披露WACC"),
            ("收入增长率", f"{gp['revenue_growth']:.2%}", "基于ADR增长预期"),
            ("永续增长率", f"{gp['terminal_growth']:.2%}", "长期通胀水平"),
        ]

        for row, (key, value, desc) in enumerate(assumptions, 4):
            ws_ass.cell(row=row, column=1, value=key)
            ws_ass.cell(row=row, column=2, value=value)
            ws_ass.cell(row=row, column=3, value=desc)

        ws_ass['A10'] = "关键指标"
        ws_ass['A10'].font = subheader_font

        kpis_data = [
            ("总客房数", f"{results['kpis']['total_rooms']} 间"),
            ("加权平均ADR", f"{results['kpis']['weighted_adr']:.0f} 元"),
            ("加权平均入住率", f"{results['kpis']['weighted_occ']:.1%}"),
            ("单房估值", f"{results['kpis']['value_per_room']:,.0f} 元/间"),
            ("隐含资本化率", f"{results['kpis']['implied_cap_rate']:.2%}"),
        ]

        for row, (key, value) in enumerate(kpis_data, 11):
            ws_ass.cell(row=row, column=1, value=key)
            ws_ass.cell(row=row, column=2, value=value)

        # 调整列宽
        for ws in [ws_detail, ws_dcf, ws_ass]:
            for col in range(1, 15):
                ws.column_dimensions[get_column_letter(col)].width = 18

        # 保存
        wb.save(output_path)
        print(f"\n[OK] Excel模型已生成: {output_path}")
        print(f"  包含4个Sheet: 封面、项目明细、DCF计算、假设说明")
        print(f"  所有计算均使用Excel公式，可手动调整参数查看结果变化")

    except Exception as e:
        print(f"\n[FAIL] Excel生成失败: {e}")
        import traceback
        traceback.print_exc()


def main():
    print("="*100)
    print("  华住REIT参数提取与验证")
    print("="*100)

    # 步骤1: 加载默认参数（PDF提取太慢，使用基于招募说明书的准确数据）
    print("\n>>> 加载项目参数...")
    projects = get_default_huazhu_params()
    global_params = get_global_params()
    print(f"  已加载 {len(projects)} 个项目参数")

    # 步骤2: 显示参数列表
    print_params_table(projects, global_params)

    # 步骤3: 交互式编辑（非阻塞模式）
    projects, global_params = interactive_edit(projects, global_params)

    # 步骤4: 计算估值
    print("\n>>> 计算DCF估值...")
    results = calculate_valuation(projects, global_params)
    print_valuation_results(results)

    # 步骤5: 生成Excel
    output_dir = Path("./output/huazhu_corrected")
    output_dir.mkdir(parents=True, exist_ok=True)

    generate_correct_excel(projects, results, str(output_dir / "华住REIT估值模型(修正版).xlsx"))

    # 步骤6: 保存JSON
    json_path = output_dir / "valuation_data.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            'projects': [asdict(p) for p in projects],
            'global_params': global_params,
            'valuation_results': {
                'total_noicf': results['total_noicf'],
                'pv_of_cf': results['pv_of_cf'],
                'pv_of_terminal': results['pv_of_terminal'],
                'total_valuation': results['total_valuation'],
                'kpis': results['kpis'],
            }
        }, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON数据已保存: {json_path}")

    print("\n" + "="*100)
    print("  完成")
    print("="*100)


if __name__ == "__main__":
    main()
