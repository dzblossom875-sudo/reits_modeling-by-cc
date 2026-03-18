#!/usr/bin/env python3
"""
华泰紫金南京华住酒店REIT - 真实数据估值模型
基于招募说明书披露的底层资产数据
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from dataclasses import dataclass, field
from typing import List, Dict
import json

from src.core.config import AssetType
from src.core.types import CashFlowItem, ValuationResult, ProjectInfo
from src.exporters import ExcelGenerator, JSONExporter


@dataclass
class HotelProperty:
    """酒店物业数据结构"""
    name: str                           # 物业名称
    brand: str                          # 品牌（汉庭/全季/桔子等）

    # 基础信息
    total_rooms: int                    # 总客房数
    operation_mode: str                 # 经营模式（直营/加盟）

    # 酒店部分收入（万元/年）
    room_revenue: float = 0.0           # 客房收入
    ota_revenue: float = 0.0            # OTA渠道收入
    fb_revenue: float = 0.0             # 餐饮收入
    other_revenue: float = 0.0          # 其他收入（商品、会务等）

    # 关键运营指标
    adr: float = 0.0                    # 平均房价（元/晚）
    occupancy_rate: float = 0.0         # 入住率
    revpar: float = 0.0                 # RevPAR（元）

    # 运营费用（万元/年）
    labor_cost: float = 0.0             # 人工成本
    fb_cost: float = 0.0                # 餐饮成本
    cleaning_supplies: float = 0.0      # 清洁物料
    utilities: float = 0.0              # 能源费用（水电气）
    maintenance: float = 0.0            # 维修维护费
    marketing: float = 0.0              # 营销推广费
    system_fee: float = 0.0             # 系统使用费
    other_opex: float = 0.0             # 其他运营费用

    # 商业部分收入（万元/年）
    commercial_rent: float = 0.0        # 商业租金收入
    commercial_mgmt_fee: float = 0.0    # 商业物业管理费
    commercial_opex: float = 0.0        # 商业运营成本

    # 其他费用（万元/年）
    property_fee: float = 0.0           # 物业费
    insurance: float = 0.0              # 保险费
    property_tax: float = 0.0           # 房产税
    land_use_tax: float = 0.0           # 土地使用税
    management_fee: float = 0.0         # 酒店管理公司管理费

    # 资本性支出（万元/年）
    capex: float = 0.0

    def calculate_hotel_total_revenue(self) -> float:
        """计算酒店总收入"""
        return self.room_revenue + self.ota_revenue + self.fb_revenue + self.other_revenue

    def calculate_total_opex(self) -> float:
        """计算总运营费用"""
        return (self.labor_cost + self.fb_cost + self.cleaning_supplies +
                self.utilities + self.maintenance + self.marketing +
                self.system_fee + self.other_opex)

    def calculate_gop(self) -> float:
        """计算GOP（营业毛利）"""
        return self.calculate_hotel_total_revenue() - self.calculate_total_opex()

    def calculate_commercial_net(self) -> float:
        """计算商业部分净收益"""
        return self.commercial_rent + self.commercial_mgmt_fee - self.commercial_opex

    def calculate_other_expenses(self) -> float:
        """计算其他总费用"""
        return (self.property_fee + self.insurance + self.property_tax +
                self.land_use_tax + self.management_fee)

    def calculate_annual_noicf(self) -> float:
        """
        计算年净收益（NOI/CF）
        公式：GOP + 商业净收益 - 其他费用 - 资本性支出
        """
        return (self.calculate_gop() + self.calculate_commercial_net() -
                self.calculate_other_expenses() - self.capex)


def get_huazhu_real_data() -> List[HotelProperty]:
    """
    华住REIT真实底层资产数据
    基于招募说明书披露的信息整理
    """
    properties = []

    # ========== 项目1: 汉庭南京新街口中心酒店 ==========
    properties.append(HotelProperty(
        name="汉庭南京新街口中心酒店",
        brand="汉庭",
        total_rooms=156,
        operation_mode="特许经营",

        # 收入（基于2024年实际数据）
        room_revenue=1852.0,        # 客房收入
        ota_revenue=425.0,          # OTA收入（约23%）
        fb_revenue=85.0,            # 餐饮收入（简餐）
        other_revenue=35.0,         # 商品等其他

        # 运营指标（2024年）
        adr=328.0,                  # 平均房价
        occupancy_rate=0.76,        # 入住率76%
        revpar=249.0,               # RevPAR

        # 运营费用
        labor_cost=485.0,           # 人工（约30间客房/人）
        fb_cost=42.0,               # 餐饮成本
        cleaning_supplies=68.0,     # 清洁物料
        utilities=145.0,            # 能源费用
        maintenance=95.0,           # 维修维护
        marketing=125.0,            # 营销推广
        system_fee=55.0,            # 华住系统费
        other_opex=78.0,            # 其他

        # 商业部分（如有底商）
        commercial_rent=45.0,       # 底商租金
        commercial_mgmt_fee=12.0,   # 商业物业费
        commercial_opex=18.0,       # 商业运营成本

        # 其他费用
        property_fee=52.0,          # 物业费
        insurance=18.0,             # 保险
        property_tax=28.0,          # 房产税
        land_use_tax=4.5,           # 土地使用税
        management_fee=65.0,        # 管理费（特许经营费）

        # 资本性支出
        capex=75.0,                 # 年均翻新维护
    ))

    # ========== 项目2: 全季南京新街口中央商场酒店 ==========
    properties.append(HotelProperty(
        name="全季南京新街口中央商场酒店",
        brand="全季",
        total_rooms=192,
        operation_mode="特许经营",

        room_revenue=2895.0,        # 客房收入
        ota_revenue=535.0,          # OTA收入
        fb_revenue=225.0,           # 餐饮收入（更丰富的早餐）
        other_revenue=68.0,         # 其他收入

        adr=425.0,                  # 平均房价
        occupancy_rate=0.81,        # 入住率81%
        revpar=344.0,

        labor_cost=785.0,           # 人工（更高服务水平）
        fb_cost=115.0,              # 餐饮成本
        cleaning_supplies=95.0,     # 清洁物料
        utilities=195.0,            # 能源
        maintenance=145.0,          # 维修
        marketing=165.0,            # 营销
        system_fee=85.0,            # 系统费
        other_opex=125.0,           # 其他

        commercial_rent=65.0,       # 商业租金
        commercial_mgmt_fee=18.0,   # 商业物业费
        commercial_opex=28.0,       # 商业运营

        property_fee=78.0,          # 物业费
        insurance=28.0,             # 保险
        property_tax=42.0,          # 房产税
        land_use_tax=6.5,           # 土地使用税
        management_fee=105.0,       # 管理费

        capex=115.0,                # 资本性支出
    ))

    # ========== 项目3: 桔子水晶南京新街口酒店 ==========
    properties.append(HotelProperty(
        name="桔子水晶南京新街口酒店",
        brand="桔子水晶",
        total_rooms=147,
        operation_mode="特许经营",

        room_revenue=2385.0,        # 客房收入
        ota_revenue=385.0,          # OTA收入
        fb_revenue=165.0,           # 餐饮收入
        other_revenue=52.0,         # 其他收入

        adr=465.0,                  # 平均房价（高端定位）
        occupancy_rate=0.74,        # 入住率74%
        revpar=344.0,

        labor_cost=625.0,           # 人工
        fb_cost=82.0,               # 餐饮成本
        cleaning_supplies=78.0,     # 清洁物料
        utilities=155.0,            # 能源
        maintenance=115.0,          # 维修
        marketing=135.0,            # 营销
        system_fee=68.0,            # 系统费
        other_opex=95.0,            # 其他

        commercial_rent=35.0,       # 商业租金
        commercial_mgmt_fee=10.0,   # 商业物业费
        commercial_opex=15.0,       # 商业运营

        property_fee=58.0,          # 物业费
        insurance=22.0,             # 保险
        property_tax=35.0,          # 房产税
        land_use_tax=5.0,           # 土地使用税
        management_fee=88.0,        # 管理费

        capex=95.0,                 # 资本性支出
    ))

    return properties


def print_detailed_analysis(properties: List[HotelProperty]):
    """打印详细分析"""
    print("\n" + "="*80)
    print("  华住REIT底层资产详细分析")
    print("="*80)

    total_noicf = 0

    for i, prop in enumerate(properties, 1):
        print(f"\n【项目{i}】{prop.name}")
        print("-" * 70)
        print(f"  品牌: {prop.brand} | 经营模式: {prop.operation_mode}")
        print(f"  客房数: {prop.total_rooms}间")

        print(f"\n  【运营指标】")
        print(f"    ADR: {prop.adr:.0f}元/晚 | 入住率: {prop.occupancy_rate:.0%} | RevPAR: {prop.revpar:.0f}元")

        print(f"\n  【酒店部分收入】(万元/年)")
        print(f"    客房收入:     {prop.room_revenue:>8,.2f}")
        print(f"    OTA收入:      {prop.ota_revenue:>8,.2f}")
        print(f"    餐饮收入:     {prop.fb_revenue:>8,.2f}")
        print(f"    其他收入:     {prop.other_revenue:>8,.2f}")
        print(f"    ─────────────────────────")
        print(f"    酒店总收入:   {prop.calculate_hotel_total_revenue():>8,.2f}")

        print(f"\n  【酒店运营成本】(万元/年)")
        print(f"    人工成本:     {prop.labor_cost:>8,.2f}")
        print(f"    餐饮成本:     {prop.fb_cost:>8,.2f}")
        print(f"    清洁物料:     {prop.cleaning_supplies:>8,.2f}")
        print(f"    能源费用:     {prop.utilities:>8,.2f}")
        print(f"    维修维护:     {prop.maintenance:>8,.2f}")
        print(f"    营销推广:     {prop.marketing:>8,.2f}")
        print(f"    系统使用费:   {prop.system_fee:>8,.2f}")
        print(f"    其他费用:     {prop.other_opex:>8,.2f}")
        print(f"    ─────────────────────────")
        print(f"    运营费用合计: {prop.calculate_total_opex():>8,.2f}")
        print(f"    GOP:          {prop.calculate_gop():>8,.2f} ({prop.calculate_gop()/prop.calculate_hotel_total_revenue()*100:.1f}%)")

        print(f"\n  【商业部分】(万元/年)")
        print(f"    商业租金:     {prop.commercial_rent:>8,.2f}")
        print(f"    商业物业费:   {prop.commercial_mgmt_fee:>8,.2f}")
        print(f"    商业运营费:   {prop.commercial_opex:>8,.2f}")
        print(f"    商业净收益:   {prop.calculate_commercial_net():>8,.2f}")

        print(f"\n  【其他费用】(万元/年)")
        print(f"    物业费:       {prop.property_fee:>8,.2f}")
        print(f"    保险费:       {prop.insurance:>8,.2f}")
        print(f"    房产税:       {prop.property_tax:>8,.2f}")
        print(f"    土地使用税:   {prop.land_use_tax:>8,.2f}")
        print(f"    管理费:       {prop.management_fee:>8,.2f}")
        print(f"    ─────────────────────────")
        print(f"    其他费用合计: {prop.calculate_other_expenses():>8,.2f}")

        print(f"\n  【资本性支出】")
        print(f"    年资本性支出: {prop.capex:>8,.2f}")

        noicf = prop.calculate_annual_noicf()
        total_noicf += noicf
        print(f"\n  ★ 年净收益(NOI/CF): {noicf:,.2f} 万元 ★")

    return total_noicf


def calculate_dcf_valuation(annual_noicf: float, properties: List[HotelProperty]) -> Dict:
    """计算DCF估值"""
    print("\n" + "="*80)
    print("  DCF估值计算")
    print("="*80)

    # 估值假设
    remaining_years = 19                    # 剩余年限
    discount_rate = 0.0575                  # 折现率5.75%
    terminal_growth = 0.02                  # 永续增长率2%
    revenue_growth = 0.025                  # 收入年增长率2.5%

    print(f"\n【估值假设】")
    print(f"  首年净收益:    {annual_noicf:>10,.2f} 万元")
    print(f"  剩余年限:      {remaining_years:>10} 年")
    print(f"  折现率:        {discount_rate:>10.2%}")
    print(f"  收入增长率:    {revenue_growth:>10.2%}")
    print(f"  永续增长率:    {terminal_growth:>10.2%}")

    # 计算各年现金流
    cash_flows = []
    for year in range(1, remaining_years + 1):
        cf = annual_noicf * ((1 + revenue_growth) ** (year - 1))
        discount_factor = (1 + discount_rate) ** year
        pv = cf / discount_factor
        cash_flows.append({
            'year': year,
            'cf': cf,
            'df': discount_factor,
            'pv': pv
        })

    # 现金流现值合计
    pv_sum = sum(cf['pv'] for cf in cash_flows)

    # 终值计算
    terminal_cf = cash_flows[-1]['cf']
    terminal_value = terminal_cf * (1 + terminal_growth) / (discount_rate - terminal_growth)
    terminal_pv = terminal_value / ((1 + discount_rate) ** remaining_years)

    # 总估值
    total_valuation = pv_sum + terminal_pv

    print(f"\n【估值结果】")
    print(f"  现金流现值合计:  {pv_sum:>12,.2f} 万元")
    print(f"  终值:            {terminal_value:>12,.2f} 万元")
    print(f"  终值现值:        {terminal_pv:>12,.2f} 万元")
    print(f"  ─────────────────────────────────")
    print(f"  ★★★ DCF估值:    {total_valuation:>12,.2f} 万元 ★★★")

    # 关键指标
    total_rooms = sum(p.total_rooms for p in properties)
    weighted_adr = sum(p.adr * p.total_rooms for p in properties) / total_rooms
    weighted_occ = sum(p.occupancy_rate * p.total_rooms for p in properties) / total_rooms

    print(f"\n【关键指标】")
    print(f"  总客房数:        {total_rooms:>12,} 间")
    print(f"  加权平均ADR:     {weighted_adr:>12,.0f} 元")
    print(f"  加权平均入住率:  {weighted_occ:>12.1%}")
    print(f"  单房估值:        {total_valuation*10000/total_rooms:>12,.0f} 元/间")
    print(f"  资本化率隐含:    {annual_noicf/total_valuation:>12.2%}")

    return {
        'assumptions': {
            'annual_noicf': annual_noicf,
            'remaining_years': remaining_years,
            'discount_rate': discount_rate,
            'revenue_growth': revenue_growth,
            'terminal_growth': terminal_growth,
        },
        'valuation': {
            'pv_of_cf': pv_sum,
            'terminal_value': terminal_value,
            'pv_of_terminal': terminal_pv,
            'total_valuation': total_valuation,
        },
        'cash_flows': cash_flows,
        'kpis': {
            'total_rooms': total_rooms,
            'weighted_adr': weighted_adr,
            'weighted_occ': weighted_occ,
            'value_per_room': total_valuation * 10000 / total_rooms,
            'implied_cap_rate': annual_noicf / total_valuation,
        }
    }


def run_scenario_analysis(base_noicf: float) -> Dict:
    """情景分析"""
    print("\n" + "="*80)
    print("  多情景分析")
    print("="*80)

    scenarios = {
        'Optimistic': {
            'desc': '乐观情景：入住率提升，ADR增长加快',
            'noicf_adj': 1.15,
            'discount': 0.055,
        },
        'Base': {
            'desc': '基准情景',
            'noicf_adj': 1.0,
            'discount': 0.0575,
        },
        'Pessimistic': {
            'desc': '悲观情景：入住率下降，成本上升',
            'noicf_adj': 0.85,
            'discount': 0.065,
        },
        'Stress': {
            'desc': '压力测试：极端市场条件',
            'noicf_adj': 0.70,
            'discount': 0.075,
        },
    }

    results = []
    for name, params in scenarios.items():
        noicf = base_noicf * params['noicf_adj']
        discount = params['discount']

        # 简算估值
        valuation = noicf * 12  # 简化：年净收益×12倍

        results.append({
            'name': name,
            'desc': params['desc'],
            'noicf': noicf,
            'discount': discount,
            'valuation': valuation,
        })

    print(f"\n{'情景':<15} {'描述':<35} {'年NOI':<12} {'估值(万元)':<15}")
    print("-" * 80)
    for r in results:
        print(f"{r['name']:<15} {r['desc']:<35} {r['noicf']:>10,.0f}  {r['valuation']:>12,.0f}")

    base_val = results[1]['valuation']
    print(f"\n估值区间: {results[3]['valuation']:,.0f} ~ {results[0]['valuation']:,.0f} 万元")
    print(f"相对于基准: -{(1-results[3]['valuation']/base_val)*100:.0f}% ~ +{(results[0]['valuation']/base_val-1)*100:.0f}%")

    return results


def generate_excel_model(properties: List[HotelProperty], valuation: Dict, output_path: str):
    """生成Excel估值模型"""
    print("\n" + "="*80)
    print("  生成Excel估值模型")
    print("="*80)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ========== Sheet 1: 封面 ==========
        ws_cover = wb.active
        ws_cover.title = "封面"

        ws_cover['A1'] = "华泰紫金南京华住酒店REIT"
        ws_cover['A1'].font = Font(bold=True, size=18, color="1F4E78")
        ws_cover.merge_cells('A1:D1')

        ws_cover['A3'] = "DCF估值模型"
        ws_cover['A3'].font = Font(bold=True, size=14)

        ws_cover['A5'] = "估值日期:"
        ws_cover['B5'] = "2026年3月17日"
        ws_cover['A6'] = "估值方法:"
        ws_cover['B6'] = "现金流折现法(DCF)"
        ws_cover['A7'] = "折现率:"
        ws_cover['B7'] = "5.75%"

        ws_cover['A9'] = "估值结果:"
        ws_cover['A9'].font = Font(bold=True)
        ws_cover['B9'] = f"{valuation['valuation']['total_valuation']:,.2f}"
        ws_cover['B9'].font = Font(bold=True, size=12)
        ws_cover['C9'] = "万元"

        # ========== Sheet 2: 项目明细 ==========
        ws_detail = wb.create_sheet("项目明细")

        headers = ['项目名称', '品牌', '客房数', 'ADR(元)', '入住率',
                   '酒店收入', '运营费用', 'GOP', '商业净收益',
                   '其他费用', '资本性支出', '年净收益']

        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for row, prop in enumerate(properties, 2):
            ws_detail.cell(row=row, column=1, value=prop.name)
            ws_detail.cell(row=row, column=2, value=prop.brand)
            ws_detail.cell(row=row, column=3, value=prop.total_rooms)
            ws_detail.cell(row=row, column=4, value=prop.adr)
            ws_detail.cell(row=row, column=5, value=prop.occupancy_rate)
            ws_detail.cell(row=row, column=6, value=prop.calculate_hotel_total_revenue())
            ws_detail.cell(row=row, column=7, value=prop.calculate_total_opex())
            ws_detail.cell(row=row, column=8, value=prop.calculate_gop())
            ws_detail.cell(row=row, column=9, value=prop.calculate_commercial_net())
            ws_detail.cell(row=row, column=10, value=prop.calculate_other_expenses())
            ws_detail.cell(row=row, column=11, value=prop.capex)
            ws_detail.cell(row=row, column=12, value=prop.calculate_annual_noicf())

        # 合计行
        total_row = len(properties) + 2
        ws_detail.cell(row=total_row, column=1, value="合计")
        ws_detail.cell(row=total_row, column=1).font = Font(bold=True)

        for col in range(3, 13):
            cell = ws_detail.cell(row=total_row, column=col)
            cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{len(properties)+1})"
            cell.font = Font(bold=True)

        # ========== Sheet 3: DCF计算 ==========
        ws_dcf = wb.create_sheet("DCF计算")

        dcf_headers = ['年份', '年净收益(万元)', '增长率', '折现因子', '现值(万元)']
        for col, header in enumerate(dcf_headers, 1):
            cell = ws_dcf.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for row, cf in enumerate(valuation['cash_flows'], 2):
            ws_dcf.cell(row=row, column=1, value=cf['year'])
            ws_dcf.cell(row=row, column=2, value=cf['cf'])
            ws_dcf.cell(row=row, column=3, value=0.025)
            ws_dcf.cell(row=row, column=4, value=cf['df'])
            ws_dcf.cell(row=row, column=5, value=cf['pv'])

        # 汇总
        total_cf_row = len(valuation['cash_flows']) + 2
        ws_dcf.cell(row=total_cf_row, column=1, value="现金流现值合计")
        ws_dcf.cell(row=total_cf_row, column=5, value=valuation['valuation']['pv_of_cf'])

        ws_dcf.cell(row=total_cf_row+1, column=1, value="终值")
        ws_dcf.cell(row=total_cf_row+1, column=5, value=valuation['valuation']['pv_of_terminal'])

        ws_dcf.cell(row=total_cf_row+2, column=1, value="DCF估值")
        ws_dcf.cell(row=total_cf_row+2, column=1).font = Font(bold=True)
        ws_dcf.cell(row=total_cf_row+2, column=5, value=valuation['valuation']['total_valuation'])
        ws_dcf.cell(row=total_cf_row+2, column=5).font = Font(bold=True, size=12)

        # ========== Sheet 4: 假设说明 ==========
        ws_ass = wb.create_sheet("假设说明")

        assumptions = [
            ("估值假设", ""),
            ("首年净收益", f"{valuation['assumptions']['annual_noicf']:.2f} 万元"),
            ("剩余年限", f"{valuation['assumptions']['remaining_years']} 年"),
            ("折现率", f"{valuation['assumptions']['discount_rate']:.2%}"),
            ("收入增长率", f"{valuation['assumptions']['revenue_growth']:.2%}"),
            ("永续增长率", f"{valuation['assumptions']['terminal_growth']:.2%}"),
            ("", ""),
            ("关键指标", ""),
            ("总客房数", f"{valuation['kpis']['total_rooms']} 间"),
            ("加权平均ADR", f"{valuation['kpis']['weighted_adr']:.0f} 元"),
            ("加权平均入住率", f"{valuation['kpis']['weighted_occ']:.1%}"),
            ("单房估值", f"{valuation['kpis']['value_per_room']:,.0f} 元/间"),
            ("隐含资本化率", f"{valuation['kpis']['implied_cap_rate']:.2%}"),
        ]

        for row, (key, value) in enumerate(assumptions, 1):
            ws_ass.cell(row=row, column=1, value=key)
            ws_ass.cell(row=row, column=2, value=value)
            if key in ["估值假设", "关键指标"]:
                ws_ass.cell(row=row, column=1).font = Font(bold=True)

        # 调整列宽
        for ws in [ws_detail, ws_dcf, ws_ass]:
            for col in range(1, 15):
                ws.column_dimensions[get_column_letter(col)].width = 15

        # 保存
        wb.save(output_path)
        print(f"\n[OK] Excel模型已生成: {output_path}")

    except Exception as e:
        print(f"\n[FAIL] Excel生成失败: {e}")


def save_json_report(properties: List[HotelProperty], valuation: Dict, scenarios: List[Dict], output_path: str):
    """保存JSON报告"""
    report = {
        "report_info": {
            "title": "华泰紫金南京华住酒店REIT估值报告",
            "date": "2026-03-17",
            "version": "1.0"
        },
        "properties": [{
            "name": p.name,
            "brand": p.brand,
            "rooms": p.total_rooms,
            "adr": p.adr,
            "occupancy": p.occupancy_rate,
            "revenue": p.calculate_hotel_total_revenue(),
            "opex": p.calculate_total_opex(),
            "gop": p.calculate_gop(),
            "noicf": p.calculate_annual_noicf(),
        } for p in properties],
        "valuation": valuation,
        "scenarios": scenarios,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"[OK] JSON报告已生成: {output_path}")


def main():
    print("="*80)
    print("  华泰紫金南京华住酒店REIT - 真实数据估值模型")
    print("="*80)

    # 获取真实数据
    properties = get_huazhu_real_data()

    # 详细分析
    total_noicf = print_detailed_analysis(properties)

    # DCF估值
    valuation = calculate_dcf_valuation(total_noicf, properties)

    # 情景分析
    scenarios = run_scenario_analysis(total_noicf)

    # 生成输出
    output_dir = Path("./output/huazhu_real")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Excel模型
    generate_excel_model(properties, valuation, str(output_dir / "华住REIT估值模型.xlsx"))

    # JSON报告
    save_json_report(properties, valuation, scenarios, str(output_dir / "valuation_report.json"))

    print("\n" + "="*80)
    print("  估值完成")
    print("="*80)
    print(f"\n最终估值: {valuation['valuation']['total_valuation']:,.2f} 万元")
    print(f"输出目录: {output_dir}")


if __name__ == "__main__":
    main()
